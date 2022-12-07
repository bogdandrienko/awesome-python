import openpyxl
# from openpyxl import load_workbook

# 1) читаю первую ячейку
# 2) читаю первую строку (+ for loop)
# 3) читаю вторую строку (+ for loop)
# 4) записываю в новый документ (с учётом перевёртывания)


workbook = openpyxl.load_workbook(filename='temp3/data1.xlsx')
# sheet_ranges = workbook['range names']
worksheet = workbook.active
print(worksheet['C2'], type(worksheet['C2']))  # <Cell 'Sheet1'.C2> <class 'openpyxl.cell.cell.Cell'>
print(worksheet['C2'].value)

print(worksheet.cell(2, 3))
print(worksheet.cell(2, 3).value)
# print(worksheet['D18'].value)

# start = 0
# stop = 10
# while start <= stop:

max_row = worksheet.max_row
max_column = worksheet.max_column

print("\n\n\n************\n\n\n")

# tuple1 = (12, 15, 17)
list1 = []
list2 = []

for number in range(1, 17+1, 1):  # 1 2 3 4 ... 10
    value1 = worksheet.cell(1, number).value
    list1.append(value1)
    value2 = worksheet[f"B{number}"].value
    list2.append(value2)
print(list1)
print(list2)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

index_row1 = 0
for row1 in list1:  # [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    index_row1 = index_row1 + 1
    new_worksheet.cell(index_row1, 1, row1)

index_row2 = 0
for row2 in list2:  # ['a', 'б', 'Bogdan', 'г', 'д', 'a', 'б', 'в', 'г', 'a', 'б', 'в', 'г', 'д', 'г', 'д', 'd']
    index_row2 = index_row2 + 1
    new_worksheet[f"B{index_row2}"] = row2

# for number in range(1, 17+1, 1):  # 1 2 3 4 ... 10
#     value1 = new_worksheet.cell(1, number).value
#     list1.append(value1)
#     value2 = new_worksheet.cell(2, number).value
#     list2.append(value2)

# new_workbook.save("temp3/new_data.xlsx")


################################################################

print("\n\n\n************\n\n\n\n\n\n\n\n\n")

workbook = openpyxl.load_workbook(filename='temp3/data1.xlsx')
worksheet = workbook.active

rows = []

for row in range(1, worksheet.max_row+1, 1):
    local_row = []
    for column in range(1, worksheet.max_column + 1, 1):
        value = worksheet.cell(row, column).value
        local_row.append(value)
    rows.append(local_row)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

for row_i, row in enumerate(rows, 1):
    for column_i, value in enumerate(row, 1):
        new_worksheet.cell(column_i, row_i, value)

new_workbook.save("temp3/new_data.xlsx")


from math import ceil, floor
from typing import Union

from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

# xlsx xls xlsm

# 1) положим их в одну папку и переименуем на латинницу
# 2) найти как прочитать их начинку (две переменные с данными)   Read
# 3) две переменные => одну переменную
# 4) создать новый документ с новыми данными                     Create
# 5) обновить документ с новыми данными                          Update


# CRUD
# Create - документа ещё нет, нужно его создать
# Read - документ уже есть, нужно прочитать
# Update - документ уже есть, нужно изменить данные
# Delete


workbook = openpyxl.load_workbook("data/data1.xlsx")
worksheet = workbook.active  # workbook["Лист 2"]

# dict1 = {"name": "Python"}
# dict1["name"]

# cella1 = worksheet["B10"].value
# print(cella1, type(cella1))
#
# cella2 = worksheet.cell(10, 2).value
# print(cella2, type(cella2))

arr1 = []
#           start, stop, step
#               range(20) # [0....19]
# for i in [2, 3, 4, 5, 1]:  #
for i in range(2, 14 + 1, 1):  # [2, 3, 4, ... 14]
    cell_value = worksheet[f"B{i}"].value
    arr1.append(cell_value)
    #
#
# print(arr1, type(arr1))

arr2 = []
# for char in ["A", "B", "C"]:  # [2, 3, 4, ... 14]
# for char in "ABCDE":  # [2, 3, 4, ... 14]
# for char in range(1, 4+1, 1):  # [2, 3, 4, ... 14]
#     # cell_value = worksheet.cell(2, char).value
#     val1 = get_column_letter(char)
#     print(val1)
#
#     cell_value = worksheet[f"{val1}2"].value
#     # cell_value = worksheet[f"{char}2"].value
#     arr2.append(cell_value)
#     #
# #
# print(arr2, type(arr2))

# max_row = worksheet.max_row
# max_column = worksheet.max_column


# arr5 = [1, 2, 3, 5]
# arr6 = [[1, [1, 2]], [1, 2], [1, 2], [1, 2]]
# todo создаём внешний пустой массив
external_array1 = []

# todo проходим циклом по строкам
for row in range(1, worksheet.max_row + 1):

    # todo создаём внутренний пустой массив
    internal_array = []

    # todo проходим циклом по столбцам
    for column in range(1, worksheet.max_column + 1):
        # todo наполняем внутренний массив значения ячеек
        internal_array.append(worksheet.cell(row, column).value)

    # todo наполняем внешний массив массивом со значением строк
    external_array1.append(internal_array)

# print(external_array1)
# for i in external_array1:
#     print(i)

workbook2 = openpyxl.load_workbook("data/data2.xlsx")
worksheet2 = workbook2.active

# todo создаём внешний пустой массив
external_array2 = []

# todo проходим циклом по столбцам
for column in range(1, worksheet2.max_column + 1):

    # todo создаём внутренний пустой массив
    internal_array2 = []

    # todo проходим циклом по строкам
    for row in range(2, worksheet2.max_row + 1):
        # todo наполняем внутренний массив значения ячеек
        internal_array2.append(worksheet2.cell(row, column).value)

    # todo наполняем внешний массив массивом со значением строк
    external_array2.append(internal_array2)

# print(external_array2)
# for i in external_array2:
#     print(i)

external_array1.extend(external_array2)
# print(external_array1)
# for i in external_array1:
#     print(i)

# todo создание новой рабочей книги в оперативной памяти
workbook3 = openpyxl.Workbook()

# todo выбор активного рабочего листа
worksheet3 = workbook3.active

# todo присвоение
# worksheet3['B3'] = 42
# worksheet3.cell(5, 3, 666)
#
# for column in range(1, 1000):
#     worksheet3.cell(3, column, 1000-column)
#
# for row in range(1, 300):
#     worksheet3.cell(row, 2, "Python")
#
# name = "Alema"
# for row in range(8, 12+1):
#     for column in range(4, 7+1):
#         worksheet3.cell(row, column, name)

print(external_array1)

external_counter = 0
# todo проходим по внешнему массиву, каждый раз берём внутренний массив
for external in external_array1:  # ['И.И.И.', 30, 600000]
    external_counter += 1

    internal_counter = 0
    # todo проходим по внутреннему массиву, каждый раз берём значение
    for internal in external:  # 'И.И.И.'  30  600000
        internal_counter += 1

        # print(f"строка: {external_counter}, колонка: {internal_counter}, значение: {internal}")

        # todo устанавливаем значени в ячейку
        worksheet3.cell(external_counter, internal_counter, internal)
#      0   1   2  3  4  !5  !6
arr = [5, "P", 2, 9, 6, 5, "P"]
for i in arr:
    index = arr.index(i)
    # print(i, index)
i_counter = 0
for i in arr:
    # print(i, i_counter)

    i_counter += 1  # increment
    #
#

# unpacking - распаковка значений
print(arr)
for index, value in enumerate(arr, 666):  # enumerate - получает значение, а возвращает индекс и значение
    print(index, value)

# [5, 'P', 2, 9, 6, 5, 'P']
# [(0, 5), (1, 'P'), (2, 2)...]

# todo сохранение
workbook3.save("data/new_data.xlsx")


def double(val: Union[int, float]) -> Union[int]:
    return round(val + 2)


res = double(2.0)
print(res)

arr9 = [1, 2, 3, 4, 5, 6]
# for i in arr9:
#     result = double(val=i)
#     print(result)

# val3 = map(double, arr9)  # lazy computing - только тогда, когда его вызывают
# print(list(val3))

for i in map(double, arr9):
    print(i)

# d = float(input("Введи диаметр"))
# l = 3.14 * d
# print(l)

r = 2 * 3.14 * 5 / 2 * 3.14

print(r)
print(floor(r))
#####################################################

data = [1, 2, 3, 4, 5, 6, 7]
workbook3 = Workbook()
worksheet3 = workbook3.active

internal_counter = 0
for value in data:
    internal_counter += 1
    worksheet3.cell(internal_counter, 1, value)

workbook3.save("data/new_data3.xlsx")

#####################################################


value = -1
while value < 0:
    value = int(input("Введите положительное число: ") + input("Введите положительное число: ") +
                input("Введите "
                      "положительное число: "))

print(value)

value = 0
while True:
    if value > 0:
        break
    else:
        value = int(input("Введите положительное число: ") + input("Введите положительное число: ") + input("Введите "
                                                                                                            "положительное число: "))

        

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, GradientFill, Alignment, Side

# workbook = Workbook()  # для создания в памяти экземпляра Workbook
# worksheet = workbook.active

workbook = openpyxl.load_workbook('users.xlsx')
worksheet = workbook.active

#            0    1
new_list = [123, 124]
print(new_list[0])

new_dict = {"first": 123, 23: 666, "23": 666}
print(new_dict["first"])

value = worksheet["A56"].value
print(value)
print(type(value))

new_set = {12, 135, 656, 12}
print(new_set)
print(type(new_set))

new_tuple = (12,)

workers = []
for row in range(1, 2150):
    coordinates = "A" + str(row)  # 1 способ сложения строк
    # print(coordinates)
    coordinates = f"A{row}"  # 2 способ сложения строк
    # print(coordinates)
    coordinates = "A{}".format(row)  # 3 способ сложения строк
    # print(coordinates)
    col = "A"
    coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
    # print(coordinates)
    # txt1 = "My name is {fname}, I'm {age}".format(fname="John", age=36)
    # txt2 = "My name is {0}, I'm {1}".format("John", 36)
    # txt3 = "My name is {}, I'm {}".format("John", 36)

    worker = []
    for col in "ABCDEF":
        coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
        value = worksheet[coordinates].value
        if value is not None:
            # print(value)
            worker.append(value)
            pass
    workers.append(worker)


# print(workers)


class Worker:
    def __init__(self, first_name: str, second_name, patronymic, id, position, category):
        self.first_name = first_name
        self.second_name = second_name
        self.patronymic = patronymic
        self.id = id
        self.position = position
        self.category = category

    def get_full_name(self):
        return f"{self.first_name} {self.second_name}"


max_row = worksheet.max_row + 1
max_column = worksheet.max_column + 1

workers = []
for row in range(2, max_row):
    worker = []
    for column in range(1, max_column):
        value = worksheet.cell(row=row, column=column).value
        worker.append(value)
    # print(worker)
    # print(type(worker))
    worker = Worker(
        first_name=worker[0],
        second_name=worker[1],
        patronymic=worker[2],
        id=worker[3],
        position=worker[4],
        category=worker[5]
    )
    workers.append(worker)
    # print(worker)
    # print(type(worker))

print(workers[2])

workbook = Workbook()  # для создания в памяти экземпляра Workbook
worksheet = workbook.active

worksheet[f"A1"] = "Фамилия Имя"
worksheet.merge_cells('B2:C6')
a1 = worksheet[f"A1"]
a1.font = Font(color="00FF6600", bold=True, sz=28)

index = 1
for worker in workers:
    index += 1
    worksheet[f"A{index}"] = worker.get_full_name()

workbook.save('new1.xlsx')

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:F4')
top_left_cell = ws['B2']
top_left_cell.value = "My Cell"
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
top_left_cell.font  = Font(b=True, color="FF0000")
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
wb.save("styled.xlsx")

########################################################################################################################

from tkinter import *  # может произойти коллизия имён (перезапись предыдущего имени последующим)
from tkinter import ttk
import tkinter
from tkinter import messagebox

import openpyxl


def cancel():
    pass


# определение(создание) функции
def get_result():
    col_from_user = col_label.get()
    print(f'колонки: {col_from_user}')
    elem_from_user = elem_label.get()
    print(f'элемент: {elem_from_user}')
    ############################################################

    file_name = 'temp/sample_example_new.xlsx'
    workbook = openpyxl.load_workbook(file_name)

    # берём активную страницу из рабочей книги
    worksheet = workbook.active

    # берём последнюю строку в excel-файле
    max_row = worksheet.max_row
    print(max_row)
    # берём последнюю колонку в excel-файле
    max_column = worksheet.max_column
    print(max_column)

    # записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
    external_array = []

    # наполнение внешнего массива
    for row in range(1, max_row + 1):
        internal_array = []

        # наполнение внутреннего массива
        for col in range(1, max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ''
                # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
                # break  # (прекратить) оператор позволяет полностью остановить цикл
            internal_array.append(value)
        # завершение наполнение внутреннего массива

        if len(internal_array) < 1:
            continue

        external_array.append(internal_array)
    # завершение наполнение внешнего массива

    print(external_array)

    # '1, 2, 6' -> [1, 2, 6]
    col_list = []
    print(col_from_user)
    print(type(col_from_user))
    for i in col_from_user.split(","):  # метод, который вызывается на строках (режет строку на массив)
        # "1, 2, 3 , 5".split(",") -> ["1", " 2", " 3 ", " 5"]
        i = str(i).strip()  # " 1" / "2 " -> "1" / "2"
        i = int(i)  # "1" -> 1
        col_list.append(i)
        print(i)
        print(type(i))
    print(col_list)

    # elem_from_user
    count = 0
    for i_new in external_array:
        for j_new in i_new:
            if elem_from_user == j_new:
                count = count + 1

    ###########################################################

    # тут мы уже должны посчитать количество вхождений
    result = count
    result_label.config(text=f"{result}")


# инициализация инстанса - создание объекта ткинтер
root = Tk()

# создание главного окна
frm = ttk.Frame(root, padding=100)
root.title("Анализ данных")
root.geometry("640x480")
frm.grid()

# результат
result_label = ttk.Label(frm, text="Количество вхождений: ")
result_label.grid(column=4, row=0)

# колонки
col_label_var = tkinter.StringVar()  # переменная, которая хранит колонки
col_label = ttk.Entry(textvariable=col_label_var)
col_label.grid(column=1, row=3)

# элемент
elem_label_var = tkinter.StringVar()  # переменная, которая хранит элемент
elem_label = ttk.Entry(textvariable=elem_label_var)
elem_label.grid(column=5, row=3)

# кнопка стоп
Button(text="stop",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=cancel,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=0, row=5)

# кнопка старт
Button(text="start",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=get_result,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=2, row=5)

root.mainloop()

str1 = "124124|14121|1411234".split(sep="|")  # -> ["124124", "14121", "1411234"]

########################################################################################################################

# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# создание "объекта" из библиотеки openpyxl
workbook = Workbook()

# берём активную страницу из рабочей книги
worksheet = workbook.active

var_list = ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]

# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]

var_range = range(1, 1000, 1)
# print(var_range)
# for n in var_range:
#     print(n)

# функция range возвращает массив чисел
# x = range(6+1)
# for n in x:
#     # print(n)
#     pass

# x = range(3, 200, 2)
# for n in x:
#     print(n)

# "кривой способ решения задачи"
# for j in "ABCD":
#     row = "1"
#     col = j
#     # записываем значение в выбранную ячейку
#     worksheet[f'{col}{row}'] = str(var_list["ABCD".index(j)])

# пока вложенный цикл не отработает полностью, вторая итерация верхнего цикла не запускается
for num in var_range:
    for name in var_list:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list.index(name) + 1)
        # записываем значение в выбранную ячейку
        worksheet[f'{col}{row}'] = str(name)

# записываем значение в выбранную (А1) ячейку
# worksheet['A1'] = 42

# сохраняем рабочую книгу в excel-файл(xlsx/xls)
workbook.save("sample_example.xlsx")

########################################################################################################################

import datetime
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


dir_name = 'dist'
files = []

# читаем все файлы из заданной директории(папки), только xlsx
for filename in os.listdir(dir_name):  # os.listdir - принимает в себя директорию откуда нужно вывести все имена файлов
    # print(filename.upper())  # делает все буквы заглавными
    # print(filename.lower())  # делает все буквы прописными
    # print(filename.capitalize())  # делает все буквы как в предложении(первая заглавная, остальные маленькие)
    # print(filename.strip())  # очищает пробелы, табуляции и знаки переноса со строки по краям

    filename1 = filename.split(sep='.')  # метод для строки, который режет её по сепаратору(набор символов)
    # print(filename)
    # print(type(filename))
    # print(filename1)
    # print(type(filename1))

    if filename1[-1] == 'xlsx' or filename1[-1] == 'xls':  # учли, что формат файла может быть устаревшим
        files.append(filename)
    else:
        pass
print(files)

data = []

for file in files:  # проходим циклом по массиву с нужными именами файлов

    # конкатенация строк или сложение строк
    file_name1 = dir_name + '/' + file
    # print(file_name1)

    try:
        workbook = openpyxl.load_workbook(file_name1)  # принимает имя файла и открывает его как рабочую книгу
    except:
        # происходит выполнение код, когда код в блоке 'try' вызвал исключение(ошибку)
        print(f'Закройте файл "{file}"!')
        continue

    worksheet = workbook.active
    max_row = worksheet.max_row
    # print(max_row)
    max_column = worksheet.max_column + 1
    # print(max_column)

    dict1 = {}  # создаём пустой словарь(тип данных ключ-значение)

    city = file.split('.')[-2].split(' ')[-1]
    # print(f"Имя файла: {city}")
    if city.isdigit():
        # пропустить итерацию цикла
        continue

    dict1["Город"] = city

    # dict1["Общий охват"] = worksheet.cell(row=6, column=3).value



    #############################################################

    # тут будет логика обработки excel

    start = str(worksheet["C6"].value).split(":")[0].split("(")[1]
    # print(f"start: {start}")
    end = str(worksheet["C6"].value).split(":")[1].split(sep=")")[0]
    # print(f"end: {end}")

    # print(tuple(worksheet[start:end]))

    oxvat = ''
    for i in tuple(worksheet[start:end]):
        for j in i:
            if j.value:
                oxvat = j.value
            else:
                print(f'ВНИМАНИЕ город {city} данные не заполнены!')
                oxvat = 0

    total = worksheet.cell(row=5, column=2).value
    if total:
        pass
    else:
        total = 0
    dict1["Всего мероприятий"] = total

    dict1["Общий охват"] = oxvat

    vstr = worksheet["D5"].value
    if not vstr:
        vstr = 0
    dict1["Всего встреч"] = vstr

    #############################################################

    # worksheet[f'{col}{row}'] = str(name)

    data.append(dict1)
    # print("\n\n\n")
print(data)

# объединяем все словари в один сводный excel-файл

new_dir_name = 'result'
# ИСКЛЮЧЕНИЯ
# Exception - исключение т.е. критическая ошибка
try:
    # происходит ПОПЫТКА выполнить какой-то код, который может вызвать ошибку

    # путь относительно текущего скрипта
    os.mkdir(dir_name + "/" + new_dir_name)
    # connection - объект в памяти, который непрерывно обменивается данными
except:
    # происходит выполнение код, когда код в блоке 'try' вызвал исключение(ошибку)
    print('Папка уже существует!')
finally:
    # происходит выполнение кода, безотносительно удачного или неудачного выполнения
    # тут нужно закрыть соединение с базой данных или файл для чтения/записи
    pass


workbook2 = Workbook()
worksheet2 = workbook2.active

titles = ["Город", 'мероприятий', 'Охват', 'Встреча']
index = 1
for title in titles:
    worksheet2[f"{get_column_letter(index)}1"] = title
    index += 1

total_vstr = 0
total_obxv = 0
total_mer = 0

extra_index = 2
for row in data:
    print(row)
    print(type(row))

    # проходим по словарю циклом for
    # print(row.values())  # возвращает из словаря значения
    # print(row.keys())  # возвращает из словаря ключи
    # print(row.items())  # возвращает из словаря ключи и значения

    # for col in row.items():
    #     print(f'key = {col[0]}')
    #     print(f'value = {col[1]}')

    index = 1
    for key, value in row.items():
        # print(f'key = {key}')
        # print(f'value = {value}')

        if key == 'Всего встреч':
            total_vstr += int(value)

        if key == 'Общий охват':
            total_obxv += int(value)

        if key == 'Всего мероприятий':
            total_mer += int(value)

        worksheet2[f'{get_column_letter(index)}{extra_index}'] = value
        index += 1
    extra_index += 1

titles = ["Итого", total_mer, total_obxv, total_vstr]
index = 1
for title in titles:
    worksheet2[f"{get_column_letter(index)}{extra_index}"] = title
    index += 1


workbook2.save(f"{dir_name}/{new_dir_name}/result.xlsx")

########################################################################################################################

# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

file_name = "temp/sample_example.xlsx"

# загружаем в память уже существующий файл на диске
workbook = openpyxl.load_workbook(file_name)

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row + 1
print(max_row)
print(type(max_row))
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column + 1
print(max_column)
print(type(max_column))

index = 0
for i in range(1, max_row):

    # # получение значения с выбранной ячейки, где row - строка, column - колонка
    # # value = worksheet.cell(row=i, column=1).value
    #
    # # получение значения с выбранной ячейки, где в квадратных скобках координаты ячейки
    # value = worksheet[f"A{i}"].value
    #
    # # if value is not None:
    # #     pass
    # # if len(str(value)) >= 2:
    # if value:
    #     print(value)
    #     print(type(value))
    #     index += 1

    for j in range(1, max_column):

        # получение значения с выбранной ячейки, где row - строка, column - колонка
        value = worksheet.cell(row=i, column=j).value
        print(f"index: {i} {j}")
        # получение значения с выбранной ячейки, где в квадратных скобках координаты ячейки
        # value = worksheet[f"A{i}"].value

        "Almaty Taraz"

        # if value is not None:
        #     pass
        # if len(str(value)) >= 2:

        #         первый                второй
        if (value == "Almaty") or (value == "Taraz"):
            print(value)
            print(type(value))
            index += 1

print(index)
# for i in range(1, max_row):
#     for j in "ABCD":
#
#
#
# for name in var_list:
#     row = num
#     # 1 -> A, 3 -> C, 26 -> Z
#     col = get_column_letter(var_list.index(name) + 1)
#     # записываем значение в выбранную ячейку
#     worksheet[f'{col}{row}'] = str(name)


light = True
electro = False

#    первый    второй
if (light) and not (electro):
    print("Правда")
else:
    print("Ложь")

# or - если хоть один правда = правда
# and - если оба правда = правда

########################################################################################################################

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# чтение
# обрабатываем данные
# записываем в исходный файл (можно исходный файл удалять)


workbook = openpyxl.load_workbook("""temp/sample_example.xlsx""")

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row
print(max_row)
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column
print(max_column)

# записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
external_array = []

# наполнение внешнего массива
for row in range(1, max_row + 1):
    internal_array = []

    # наполнение внутреннего массива
    for col in range(1, max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ''
            # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
            # break  # (прекратить) оператор позволяет полностью остановить цикл
        internal_array.append(value)
    # завершение наполнение внутреннего массива

    if len(internal_array) < 1:
        continue

    external_array.append(internal_array)
# завершение наполнение внешнего массива

print(external_array)

# index = 0
# while True:
#     index += 1
#     if index > 50:
#         break # (прекратить) оператор позволяет полностью остановить цикл

workbook_new = Workbook()
worksheet_new = workbook_new.active

# col_count = external_array[0]
print(
    f"col_count: {external_array}")  # [['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''], ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''],]
print(f"col_count: {external_array[0]}")  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
print(f"col_count: {external_array[0][1]}")  # 'Almaty'
print(f"col_count: {external_array[0][1][2:-2:1]}")  # 'ma'


def function_len_array(array):  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
    return len(array)  # 5


# -2 -1 0 1 2

for row in range(0, function_len_array(external_array)):  # [0, 1, 2, 3, 4, ..., 1007]
    # print(f"col_count: {len(external_array[row])}")
    for col in range(0, function_len_array(external_array[row])):
        # [0, 1, 2, 3, 4, 5]  # external_array[row] == ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']

        worksheet[f'{get_column_letter(col + 1)}{row + 1}'] = external_array[row][col]
        if row == 0:
            worksheet[f'{get_column_letter(col + 1)}{row + 1}'].font = Font(bold=True)
        pass
    # print(external_array[row])
    pass

# for char in 'ABCDE':
# worksheet_new['A1'].font = Font(bold=True)
# worksheet_new['B1'].font = Font(bold=True)
# worksheet_new['C1'].font = Font(bold=True)
# worksheet_new['D1'].font = Font(bold=True)
# worksheet_new['E1'].font = Font(bold=True)


workbook.save('temp/sample_example_new.xlsx')

wb = Workbook()
ws = wb.active
ws['B2'] = "Hello1"
ws['B3'] = "Hello"
ws['B2'].font = Font(bold=True)
ws['B3'].font = Font(bold=True)
wb.save("temp/BoldDemo.xlsx")

# читаем первый файл, второй, третий... файлы
# складываем значение в каждом уникальном массиве
# создаём новый файл, который содержит в себе финальное значение (+ форматирование)
# как получить все имена файлов в папке

# форматирование: жирный, курсив, размер, capitalize, фон, границы, подчёркивание, формулы, объединение, подбор ширины

########################################################################################################################

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

dir_name = "home_work"
files = []
for filename in os.listdir(dir_name):
    files.append(filename)
data = []
file_name = []
for file in files:
    file_name1 = dir_name + "/" + file
    # file_name2.append(file_name1)
# print(file_name2)
    # print(type(file_name1))
    worbook = openpyxl.load_workbook(file_name1)
    worksheet = worbook.active
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    # start = str(worksheet["C3"].value)
    # print(start)
    #
    external_array = []
    for row in range(1, max_row+1):
        internal_array = []
        for col in range(1, max_column+1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ""
            internal_array.append(value)
            # if len(internal_array) < 1:
            #     continue
        external_array.append(internal_array)
    # print(external_array)
    data.append(external_array)
print(data)


arr1 = [x[0] for x in data[0]]
print(arr1)
arr2 = [x[1] for x in data[1]]
print(arr2)
arr3 = [x[2] for x in data[2]]
print(arr3)

len_max = len(arr1)
if len(arr2) > len_max:
    len_max = len(arr2)
if len(arr3) > len_max:
    len_max = len(arr3)
print(len_max)

new_arr = []
for i in range(0, len_max):
    try:
        elem1 = arr1[i]
    except Exception as error:
        elem1 = ''
    try:
        elem2 = arr2[i]
    except Exception as error:
        elem2 = ''
    try:
        elem3 = arr3[i]
    except Exception as error:
        elem3 = ''

    new_arr.append([elem1, elem2, elem3])
print(new_arr)

########################################################################################################################

import openpyxl
from openpyxl.utils import get_column_letter


class Calc:
    def __init__(self, value1: float, value2: float):
        self.value1 = value1
        self.value2 = value2

        self.sum = value1 + value2

    def sum(self):
        return self.sum

    def multiply(self):
        return self.value1 * self.value2

    @staticmethod
    def static_multiply(value1, value2):
        return value1 * value2


obj = Calc(12, 1.5)  # __init__
print(obj.multiply())  # multiply

print(Calc.static_multiply(1.5, 20))


class MyClass:
    def __init__(self, name="_1", index=1, value="_A1"):
        self.name = name
        self.index = index
        self.value = value

    def get_values(self):
        return [self.name, self.index, self.value]


records = []
for i in range(1, 1000 + 1):
    records.append(
        MyClass(
            name=f"_{i}",
            index=i,
            value=f"_A{i}",
        )
    )
print(records)

for i in records:
    print(i.get_values())

# Read all data from .xlsx
book = openpyxl.load_workbook("excel_data/tab.xlsx")
sheet = book.active
global_rows = []
for row in range(2, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)


class Record:
    def __init__(self, row: list):
        self.tovar = str(row[0])
        self.gruppa = str(row[1])
        self.postavshick = str(row[2])
        date_post = str(row[3]).split(" ")[0]  # '2010-12-4'
        year = date_post.split('-')[0]
        month = int(date_post.split('-')[1])  # '09' -> 9
        if int(month) < 10:
            month = f"0{month}"
        day = int(date_post.split('-')[2])
        if int(day) < 10:
            day = f"0{day}"
        self.date_post = str(f"{year}-{month}-{day}")
        self.region = str(row[4])
        try:
            self.prodashi = int(row[5])
        except Exception as error:
            print(error)
            self.prodashi = 0
        try:
            self.sbit = int(row[6])
        except Exception as error:
            print(error)
            self.sbit = 0
        pribil = str(row[7])
        if pribil.lower() == "да":
            self.pribil = True
        else:
            self.pribil = False


book_new = openpyxl.Workbook()
sheet_new = book_new.active
records = []
for row in global_rows:
    records.append(Record(row=row))

print(records)

row_index = 2
for row in records:
    sheet_new[f"A{row_index}"] = str(row.tovar)
    sheet_new[f"B{row_index}"] = str(row.prodashi)
    sheet_new[f"C{row_index}"] = str(row.postavshick)
    sheet_new[f"D{row_index}"] = str(row.pribil)
    row_index += 1
book_new.save('excel_data/new_new_table.xlsx')

########################################################################################################################

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
import openpyxl
import datetime
import random

# TODO CREATE AND WRITE ##################################################################

workbook = Workbook()  # создание инстанса - экземпляра класса
worksheet = workbook.active  # активация 1 рабочего листа
arr1 = [1, 151513, 12, 66, "rgwrg", datetime.datetime.now()]
arr2 = [1, 5, 12, 1124124, "rgwrg", datetime.datetime.now()]
arr3 = [1, 1241, 12, 66, "rgwrg", datetime.datetime.now()]
worksheet['B2'] = 666  # присвоение ячейки с координатами значения

# for item in set(arr1):
#     worksheet[f'A{arr1.index(item)+1}'] = item

# for elem in "ABCD":
#     index = "ABCD".index(elem)
#     worksheet[f'{elem}1'] = arr1[index]
#     worksheet[f'{elem}2'] = arr2[index]
#     worksheet[f'{elem}3'] = arr3[index]

words = ["Мяч", "Столб", "Тетрадь", "openpyxl", "worksheet", "Worksheet"]

massiv1 = [x for x in range(1, 5 + 1)]
massiv2 = [random.choice(words) for x in range(1, 5 + 1)]
massiv_s_massivami = [massiv1, massiv2, massiv1, [random.choice(words) for x in range(1, 5 + 1)]]
# print(list1)
# print(list2)
# print(random.choice(words))
# print(massiv_s_massivami)

# for row_i in range(0, len(massiv_s_massivami)):
#     massiv = massiv_s_massivami[row_i]
#
#     for col_i in range(0, len(massiv)):
#         znachenie = massiv[col_i]
#
#         print(f"row: {row_i+1}, col: {col_i+1} value: {znachenie}")
#         worksheet.cell(row=row_i+1, column=col_i+1, value=znachenie)

row_i = 0
for massiv in massiv_s_massivami:
    row_i += 1

    col_i = 0
    for znachenie in massiv:
        col_i += 1

        # print(f"row: {row_i}, col: {col_i} value: {znachenie}")
        worksheet.cell(row=row_i, column=col_i, value=znachenie)


def get_password(length: int, elem="") -> str:
    if length <= 0 or len(elem) <= 0:
        raise Exception
    password = ""
    while len(password) < length:
        password += random.choice(elem)
    return password


# print(get_password(length=8, elem="ABNFGKFGOIWNPV1234567890!=-"))

# workbook.save("example.xlsx")  # сохранение рабочей книги

# TODO READ ##################################################################

# workbook = openpyxl.load_workbook(filename="data.xlsx")
# worksheet = workbook.active

# val1 = worksheet['E1'].value
# print(val1)
# print(type(val1))

# external_matrix = []
# for i in range(1, 4+1):
#     internal_matrix = []
#     for j in range(1, 4+1):
#         # internal_matrix.append(worksheet[f'{get_column_letter(j)}{i}'].value)
#         internal_matrix.append(worksheet.cell(row=i, column=j).value)
#     external_matrix.append(internal_matrix)

# print(external_matrix)

# row1 = [worksheet.cell(row=x, column=1).value for x in range(1, 10)]
# print(row1)

# worksheet['A1'] = 111
# worksheet.cell(row=1, column=1, value=111)

# TODO READ AND UPDATE ##################################################################

workbook = Workbook()
# workbook.create_sheet()
worksheet = workbook.active
font = Font(name='Tahoma', size=16, bold=True,
            italic=False, vertAlign=None, underline='none',
            strike=False, color='FF0000FF')
worksheet['B2'].font = font
worksheet['B2'] = 'Python'

workbook.save("data.xlsx")

# workbook = openpyxl.load_workbook('document.xlsx')
# workbook.template = True
# workbook.save('document_template.xltx')


########################################################################################################################

import openpyxl
import json

# workbook = openpyxl.load_workbook("data1.xlsx")
# worksheet = workbook.active
# sell1 = worksheet["B2"].value
# print(sell1)
# list1 = []
# for i in "ABCD":
#     list1.append(worksheet[f"{i}2"].value)
# print(list1)
# new_workbook = openpyxl.Workbook()
# new_worksheet = new_workbook.active
# index = 1
# for j in list1:
#     new_worksheet[f"B{index}"] = j
#     index+=1
#
# new_worksheet["C3"] = 666
# new_workbook.save("data2.xlsx")


# workbook = openpyxl.load_workbook("data1.xlsx")
# worksheet = workbook.active
# sell1 = worksheet["B2"].value
# print(sell1)
#
# list_external =[]
# for j in range(1, 5):
#     list_internal = []
#     for i in "ABCD":
#         list_internal.append(worksheet[f"{i}{j}"].value)
#     list_external.append(list_internal)
# print(list_external)
#
# key1 = list_external[0][0]
# key2 = list_external[0][1]
# key3 = list_external[0][2]
# key4 = list_external[0][3]
# print(key1, key2, key3, key4)
#
#
#
#
# dict1 = {key1: list_external[1][0], key2: list_external[1][1], key3: list_external[1][2], key4: list_external[1][3]}
# print(dict1)
# with open("data1.json", mode="w") as file:
#     json.dump(dict1, file)

id = -1
name = 3
surname = 3
age = 4
salary = 5

########################################################################################################################


########################################################################################################################
