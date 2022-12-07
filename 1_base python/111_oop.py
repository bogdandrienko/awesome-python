# пустой класс
# создание экземпляра класса
# магические методы класса
# методы класса
# статические методы класса
# переменная класса и переменная экземпляра класса
# публичные, защищённые  и приватные переменные
# одиночное наследование
# несколько поочерёдных одиночных наследований
# множественное наследование



import openpyxl
from openpyxl import Workbook

######################################################################################################################################################################


class Mother:
    val1 = 12

    def __init__(self, val1, name="Мама") -> None:
        self.name = name
        self.val = val1
        self._val = val1  # защищённый
        self.__val2 = val1 + 5  # приватный

    def get_value2(self):
        return 10

    def __str__(self):
        return self.name


class Father:
    val2 = 13

    def __init__(self, val1) -> None:
        self.val1 = val1

    def get_value(self):
        return 8

    def __str__(self):
        return self.val1


class Child(Mother, Father):

    def __init__(self, val1) -> None:
        super().__init__(val1)

    def get_value1(self):
        return 5


print(Mother.val1)
a1 = Mother(10)
print(a1.val)
# print(a1._Mother__val2)

ch1 = Child(10)
print(ch1.get_value())





####################################################################################################################################################################



# класс, который позволит считать периметр и площадь


class MyClass:
    def __init__(self, a, b, name='квадрат'):
        self.name = name
        self.a = int(a)
        if isinstance(b, str):
            b = int(b)
        self.b = b

    def get_perimeter(self):  # метод внутри класса, который работает с внутренними параметрами
        return self.a + self.b

    @staticmethod
    def get_perimeter_static(a, b):  # метод внутри класса, который работает с внутренними параметрами
        return a + b

    def get_square(self):  # метод внутри класса, который работает с внутренними параметрами
        return self.a * self.b

    @staticmethod
    def get_square_static(a, b):  # метод внутри класса, который работает с внутренними параметрами
        return a * b


# oop Калькулятор

class MyCalculator:
    def __init__(self, val1: float, val2: float):
        self.val1 = val1
        self.val2 = val2

    def sum1(self):
        return self.val1 + self.val2

    @staticmethod
    def sum1_static(val1: float, val2: float):
        print(val1 + val2)


if __name__ == "__main__":
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = '12'
    workbook.save('new.xlsx')

    print(workbook)
    print(type(workbook))

    print(MyClass)
    myclass = MyClass(name="Прямоугольник", b=15, a=15)  # создание экземпляра класса
    print(myclass)
    print(type(myclass))
    print(myclass.a)
    print(myclass.b)
    print(myclass.name)
    print(myclass.get_perimeter())
    print(myclass.get_square())
    print(MyClass.get_perimeter_static(17, 18))

    MyCalculator.sum1_static(20, 30.0)
    value = MyCalculator(30, 21.0).sum1()
    print(value)
else:
    pass
    
###########################################################################
# class Meter():
#     def __init__(self, dev):
#         self.dev = dev
#     def __enter__(self):
#         #ttysetattr etc goes here before opening and returning the file object
#         self.fd = open(self.dev, MODE)
#         return self
#     def __exit__(self, type, value, traceback):
#         #Exception handling here
#         close(self.fd)
#
# meter = Meter('dev/tty0')
# with meter as m:
#     #here you work with the file object.
#     m.fd.read()
import datetime


# Процедурное программирование               Паскаль
# Функциональное программирование            F#
# Объектно-ориентированное программирование  C#, Делфи, Python, c++


#                     объект - видимость, рендеринг, коллизии (физики)
#                     техника - скорость, масса, может с ними взаимодействовать
#         сухопутные             водоплавающие  - точки, где можно перемещаться
#  Велосипед Машина Мотоцикл  гидроскутер  Лодка     - текстуры, цвет


# print(type(10))
# print(type(True))
# print(type([True, 12]))

class Obj:
    is_start_rendering1 = True
    visible = False  # публичное свойство экземпляра класса
    _visible = True  # защищённое свойство экземпляра класса
    __visible = True  # приватное свойство экземпляра класса

    def __init__(self, mass=10.0, is_start_rendering=True):  # магический метод(функция)
        self.mass = mass
        self.is_start_rendering1 = True
        self.is_start_rendering = is_start_rendering
        self.mass2 = 12
        self.__mass3 = 12

    def get_visibility(self) -> bool:  # метод
        """
        Return state of is visibile obj
        :return: bool
        """
        return self._visible

    def set_visibility(self, is_checked: bool) -> None:
        self._visible = is_checked
        print(self.mass2)

    def get_mass(self):
        return self.mass

    # def switch_visibility(self) -> None:  # метод
    #     """
    #     Этот метод должен переключать видимость
    #     :return: None
    #     """
    #     self._visible = not self._visible

    # def switch_and_return_visibility(self) -> bool:  # метод
    #     """
    #     Этот метод должен переключать видимость
    #     :return: None
    #     """
    #     self._visible = not self._visible
    #     return self._visible

    # def print_visibility(self) -> None:
    #     print(self._visible)


class Vehicle(Obj):
    def __init__(self, speed: float):
        super().__init__(mass=100)
        self.speed = speed
        self.mass = 200

    def get_parent_mass(self):
        return super().get_mass()

    # как получить значение такого атрибута, но от родителя
    # использование  public || private || protected
    # множественное наследование и "проблема ромба"
    # бинарное дерево


class Some:
    pass


som1 = Some()

veh1 = Vehicle(speed=50.2)
print(veh1.mass)
print(veh1.get_parent_mass())


#
# print(Obj.visible)
# print(Obj._visible)
# # print(Obj.__visible)
#
# new_obj_1 = Obj()  # создание экземпляра
# print(new_obj_1)
# print(new_obj_1.visible)
# print(new_obj_1._visible)
# # print(new_obj_1.__visible)
# print(type(new_obj_1))

# new_obj_1 = Obj()
#
# print(new_obj_1.get_visibility())
#
# # new_obj_1.switch_visibility()
#
# print(new_obj_1.get_visibility())
#
# new_obj_1.set_visibility(is_checked=not new_obj_1.get_visibility())
#
# print(new_obj_1.get_visibility())


class Parall:
    def __init__(self, side1: int, side2: int):
        self.side1 = side1
        self.side2 = side2
        self.side3 = side1 / 3 * side2

    def check_is_square(self) -> bool:
        # if self.side1 == self.side2:
        #     return True
        # else:
        #     return False
        return self.side1 == self.side2

    def get_perimeter(self) -> int:
        return (self.side1 + self.side2) * 2

    def get_square(self) -> int:
        if self.check_is_square():
            return self.side1 ** 2  # квадрат
        else:
            return self.side1 * self.side2  # многоугольник

    def __add__(self, other) -> int:  # +
        if isinstance(other, Parall):
            return self.get_perimeter() + other.get_perimeter()
        else:
            raise TypeError

    def sum_self(self, other) -> int:  # +
        if isinstance(other, Parall):
            return self.get_perimeter() * other.get_perimeter()
        else:
            raise TypeError


parall1 = Parall(side1=2, side2=3)
parall2 = Parall(side1=30, side2=30)
parall3 = Parall(side1=50, side2=30)


# print(parall1 + 10)

# list1 = [Parall(side1=x + 1, side2=x + 2) for x in range(1, 101)]
# sum1 = 0
# print(list1)
# for i in list1:
#     sum1 += i.get_square()
# print(sum1)
# print(parall1.get_perimeter())

# if parall1.check_is_square():
#     print('Это квадрат!')
# else:
#     print('Это прямоугольник!')
#
# if parall2.check_is_square():
#     print('Это квадрат!')
# else:
#     print('Это прямоугольник!')


class HelpPython:
    class Time:
        @staticmethod
        def get_current_time(timezone: int, formatting="23:59:59"):
            import datetime
            datetime = datetime.datetime.now()

            match formatting:
                case "23:59:59":
                    return datetime.strftime('%H:%M:%S')
                case "23:59:59.999":
                    return datetime.strftime('%H:%M:%S ...')
                case "23:59":
                    return datetime.strftime('%H:%M')
                case _:
                    return datetime

        @staticmethod
        def get_different_times_in_seconds(datetime1: datetime.datetime, datetime2: datetime.datetime) -> int:
            # datetime1 - datetime2
            return 0

    class Variable:
        class Types:
            @staticmethod
            def example():
                value1 = 10  # - 100
                value2 = 10.0  # -10.6
                value3 = True  # False
                value4 = ""
                value5 = "10"
                value6 = "Hello"
                str1 = b'\x01\x02\x03\x04\x05'
                str2 = "\n \t  I'am "
                str3 = """\n \t  I'am """
                str4 = '''\n \t  I'am '''
                str5 = '\n \t \\ I\'am '
                str6 = r"\n\tC:\Users\bogdan\Desktop\Django\Курс лекций Django1\Python"

        @staticmethod
        def convert_to_string(value) -> str:
            # return str(value)
            # return str(value).format(1)
            return f"{value}"

        @staticmethod
        def get_type_variable(var) -> type:
            return type(var)

        @staticmethod
        def get_value_from_dict(dictionary: dict, value: str, default_value, is_logging=False):
            try:
                # return dictionary.get("username", None)  # не вызывает исключения при отсутвии ключа
                return dictionary[value]
            except Exception as error:
                if is_logging:
                    print(f"Ошибка: {error}")
                return default_value

    class Arifmetic:
        @staticmethod
        def example():
            import math  # много математики

            val1 = -10
            val2 = 15

            result = math.sqrt(16)  # int(math.sqrt(16))
            print(type(result > 3.9))

            print(result > 3.9)

            abs(val1)

            print(10 * 4)
            print(10 ** 4)
            print(10 / 5)
            print(10 // 5)  # int(10 / 5) round(10 / 3, 2) 3.33

            print(16 ** 0.5)

            print(7 % 2)  # 1
            print(6 % 2 == 0)  # чётное

    class ExceptionHelps:
        class MyException1(Exception):
            def __init__(self, exception_text: str, sector: str, level_error: int, critical: bool):
                self.exception_text = exception_text
                self.datetime = datetime.datetime.now()
                self.sector = sector
                self.level_error = level_error
                self.critical = critical

            def print_error(self):
                print(f"{self.sector}: {self.exception_text} {self.datetime}")


dict1 = {"name": "Python"}
print(HelpPython.Variable.get_value_from_dict(dictionary=dict1, value="name", default_value="C++", is_logging=False))


class MyException(Exception):
    def __init__(self, exception_text: str, sector: str, level_error: int, critical: bool):
        self.exception_text = exception_text
        self.datetime = datetime.datetime.now()
        self.sector = sector
        self.level_error = level_error
        self.critical = critical

    def print_error(self):
        print(f"{self.sector}: {self.exception_text} {self.datetime}")


def test1(a, b):
    try:
        if b == 0:
            raise HelpPython.ExceptionHelps.MyException1(exception_text="кто-то накосячил с кодом",
                                                         sector="test1 Вычислительное ядро",
                                                         level_error=3, critical=False)
        c = a / b
        return c
    except HelpPython.ExceptionHelps.MyException1 as error:
        print(f"MY {error}")
        print(error.sector)
        print(error.print_error())
        return a
    except Exception as error:
        print(f"123124124 {error}")
        return None


print(test1(9, 0))


class BinaryTree:
    def __init__(self, data: int):
        self.left = None
        self.right = None
        self.root = data

    def get_data(self):
        return self.root

    def insert_new_data(self, data: int):
        if self.root:
            if data < self.root:

                # левый
                if self.left is not None:
                    self.left.insert_new_data(data)
                else:
                    self.left = BinaryTree(data)
                # левый

            elif data > self.root:

                # правый
                if self.right is not None:
                    self.right.insert_new_data(data)
                else:
                    self.right = BinaryTree(data)
                # правый

            else:
                print('Значение повторяется')

        else:
            self.root = data

    def print_all_edges(self):
        print(self.root)
        print(self.left)
        print(self.right)


tree1 = BinaryTree(1)  # создание экземляра класса - создание объекта
tree1.insert_new_data(2)
tree1.insert_new_data(3)
tree1.insert_new_data(9)
tree1.insert_new_data(6)
tree1.insert_new_data(11)
tree1.insert_new_data(15)
print(f"root: {tree1.root}")
print(f"root: {tree1.right.root}")

print(f"root: {tree1.right.right.root}")

print(f"root: {tree1.right.right.right.root}")

print(f"root: {tree1.right.right.right.right.root}")
print(f"root: {tree1.right.right.right.left.root}")

print(f"root: {tree1.right.right.right.right.right.root}")

# class BinaryTree2:
#     def __init__(self, data: int):
#         self.left = None
#         self.right = None
#         self.root = data
#
#
# tree1 = BinaryTree2(1)  # создание экземляра класса - создание объекта
# tree2 = BinaryTree2(2)  # присвоение в левую ветку нового объекта
# tree1.left = tree2  # присвоение в правую ветку нового объекта
# tree1.right = BinaryTree2(3)  # присвоение в правую ветку нового объекта
# tree1.right.right = BinaryTree2(6)
# print(tree1)
# print(f"root: {tree1.root}")
# print(f"root: {tree1.left.root}")

index = 0
def substr(val: int):
    global index
    index += 1
    print(f'вход номер {index}, значение {val}')

    if val == 1:
        return val
    else:
        return substr(val//2)

print(substr(12))

# CreateReadUpdateDelete
# CRUD

# Date and time is: 2021-07-03 16:21:12.357246
# Timestamp is: 1625309472.357246
Footer
© 2022 GitHub, Inc.
Footer navigation
Terms
Privacy
Security
Status
Docs
Contact GitHub
Pricing
API
Training
Blog
About

###########################################################################################


class ObjectOrientedProgrammingClass:
    @staticmethod
    def example_worker():
        class Worker:
            """
            Класс, который содержит в себе работника, со значениями по строке
            """

            def __init__(self, A_1='', B_1='', C_1='', D_1='', E_1='', F_1='', G_1='', H_1='', M_1=''):
                # Подразделение
                self.A_1 = A_1
                # Цех или Служба
                self.B_1 = B_1
                # Отдел или участок
                self.C_1 = C_1
                # Фамилия
                self.D_1 = D_1
                # Имя
                self.E_1 = E_1
                # Отчество
                self.F_1 = F_1
                # Табельный №
                self.G_1 = G_1
                # Категория
                self.H_1 = H_1
                # Пол
                self.M_1 = M_1

            def print_worker(self):
                """
                """
                print(
                    f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

            def get_worker_value(self, index):
                """
                :param index:
                :return:
                """
                value_ = list(
                    (self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
                return value_[index]

            def get_worker_id(self):
                """
                :return:
                """
                return self.G_1

    @staticmethod
    def example_objects():
        class Actions:
            def __init__(self):
                pass

            @staticmethod
            def action(first_value, second_value):
                return first_value + second_value

        print(Actions.action(10, 30))

        class Circle:
            pi = 3.14

            def __init__(self, radius=1):
                self.radius = radius
                self.circle_reference = 2 * self.pi * self.radius

            def get_area(self):
                return self.pi * (self.radius ** 2)

            def get_circle_reference(self):
                return 2 * self.pi * self.radius

        circle_1 = Circle(4)
        print(circle_1.get_area())
        print(circle_1.circle_reference)
        print(circle_1.get_circle_reference())

        # Классы
        class Car:
            wheels_number = 4

            def __init__(self, name, color, year, is_crashed):
                self.name = name
                self.color = color
                self.year = year
                self.is_crashed = is_crashed

        mazda_car = Car(name="Mazda CX7", color="red", year=2017, is_crashed=True)
        print(mazda_car.name, mazda_car.is_crashed, mazda_car.wheels_number)
        bmw_car = Car(name="Mazda", color="black", year=2019, is_crashed=False)
        print(bmw_car.name, bmw_car.is_crashed, bmw_car.wheels_number)
        print(Car.wheels_number * 3)

#####################################################################################################

import openpyxl


class Tree:
    def __init__(self, age: float, name="Дуб"):  # инициализатор
        self.age = age
        self.name = name

        self.value1 = 10  # открытое свойство
        self._value2 = 10  # protected свойство
        self.__value3 = 10  # private свойство

    # def __new__(cls, *args, **kwargs):  # конструктор
    #     pass

    def __str__(self):  # строковое представление объекта
        return self.name

    def increase(self):
        value = self.__value3
        self.age += 1
        return self.age

    @staticmethod
    def increase_static(age: float):
        return age + 1.0


tree1 = Tree(15.4, "Сосна")


# print(tree1)
# print(type(tree1))
# print(tree1.age)
# print(tree1.increase())
# print(tree1.age)
# print(tree1._value2)
#
# print(Tree.increase_static(18.9))


class Dub(Tree):
    def __init__(self, age: float, name: str):
        super().__init__(age, name)

    def get_name(self):
        return self.name


tree2 = Dub(15.4, "Сосна")


# print(tree2.age)
# print(tree2.increase())
# print(tree2.age)
# print(tree2._value2)

#
#
#
#
#                 Транспортные средства
#            Сухопутные               Морские
#    машины       мотоциклы    подводные    наводные
# электро


class Transport:
    def __init__(self, name, mass, motor, price, speed):
        self.name = name
        self._mass = mass
        self.motor = motor
        self.price = price
        self.speed = speed  # публичная - видна во всех случаях

        self._multiplayer = 12  # защищённая - предупреждает, при попытке её извлечь вне собственного контекста
        self.__multiplayer = 10  # приватная - невидима везде, кроме собственного контекста

    def drive(self):
        return self._mass / self.motor

    def get_speed(self):
        return self.speed


Transport1 = Transport("Трактор", 2000, 400, 5000, 30)

print(Transport1)
print(type(Transport1))
print(Transport1.drive())
print(Transport1.speed)
print(Transport1._mass)
print(Transport1._multiplayer)


class Water(Transport):
    def __init__(self, speed, *args):
        super().__init__(*args)

        self.speed1 = speed

    def drive(self):
        return super().drive() / 0.85

    def get_old_speed(self):
        return super().get_speed()


Transport2 = Water(1000, "Катамаран", 500, 20, 700, 15)

print(Transport2)
print(type(Transport2))
print(Transport2.drive())
print(Transport2.speed)
print(Transport2.speed1)


class SubWater(Water):
    def __init__(self, *args, **kwargs):  # args - позиционные - кортеж, kwargs - именные - словарь
        super().__init__(*args, **kwargs)

    def drive(self):
        return super().drive() * 1.5


Transport3 = Water(333333, "Подлодка", 333, 33, 3333, 3)

dict1 = {"speed": 12}
val1 = dict1.get("speed", "")
val2 = dict1["speed"]

print(Transport3)
print(type(Transport3))
print(Transport3.drive())
print(Transport3.speed)
print(Transport3._mass)
print(Transport3.speed1)
print(Transport3.get_speed())

for key, _ in {"speed": 12, "title": "Car"}.items():  # -> ((key, value), (key, value))
    print(f"{key}")

wb = openpyxl.Workbook()


import openpyxl


class Tree:
    def __init__(self, age: float, name="Дуб"):  # инициализатор
        self.age = age
        self.name = name

        self.value1 = 10  # открытое свойство
        self._value2 = 10  # protected свойство
        self.__value3 = 10  # private свойство

    # def __new__(cls, *args, **kwargs):  # конструктор
    #     pass

    def __str__(self):  # строковое представление объекта
        return self.name

    def increase(self):
        value = self.__value3
        self.age += 1
        return self.age

    @staticmethod
    def increase_static(age: float):
        return age + 1.0


tree1 = Tree(15.4, "Сосна")


# print(tree1)
# print(type(tree1))
# print(tree1.age)
# print(tree1.increase())
# print(tree1.age)
# print(tree1._value2)
#
# print(Tree.increase_static(18.9))


class Dub(Tree):
    def __init__(self, age: float, name: str):
        super().__init__(age, name)

    def get_name(self):
        return self.name


tree2 = Dub(15.4, "Сосна")


# print(tree2.age)
# print(tree2.increase())
# print(tree2.age)
# print(tree2._value2)

#
#
#
#
#                 Транспортные средства
#            Сухопутные               Морские
#    машины       мотоциклы    подводные    наводные
# электро


class Transport:
    def __init__(self, name, mass, motor, price, speed):
        self.name = name
        self._mass = mass
        self.motor = motor
        self.price = price
        self.speed = speed  # публичная - видна во всех случаях

        self._multiplayer = 12  # защищённая - предупреждает, при попытке её извлечь вне собственного контекста
        self.__multiplayer = 10  # приватная - невидима везде, кроме собственного контекста

    def drive(self):
        return self._mass / self.motor

    def get_speed(self):
        return self.speed


Transport1 = Transport("Трактор", 2000, 400, 5000, 30)

print(Transport1)
print(type(Transport1))
print(Transport1.drive())
print(Transport1.speed)
print(Transport1._mass)
print(Transport1._multiplayer)


class Water(Transport):
    def __init__(self, speed, *args):
        super().__init__(*args)

        self.speed1 = speed

    def drive(self):
        return super().drive() / 0.85

    def get_old_speed(self):
        return super().get_speed()


Transport2 = Water(1000, "Катамаран", 500, 20, 700, 15)

print(Transport2)
print(type(Transport2))
print(Transport2.drive())
print(Transport2.speed)
print(Transport2.speed1)


class SubWater(Water):
    def __init__(self, *args, **kwargs):  # args - позиционные - кортеж, kwargs - именные - словарь
        super().__init__(*args, **kwargs)

    def drive(self):
        return super().drive() * 1.5


Transport3 = Water(333333, "Подлодка", 333, 33, 3333, 3)

dict1 = {"speed": 12}
val1 = dict1.get("speed", "")
val2 = dict1["speed"]

print(Transport3)
print(type(Transport3))
print(Transport3.drive())
print(Transport3.speed)
print(Transport3._mass)
print(Transport3.speed1)
print(Transport3.get_speed())

for key, _ in {"speed": 12, "title": "Car"}.items():  # -> ((key, value), (key, value))
    print(f"{key}")

wb = openpyxl.Workbook()

import openpyxl
from openpyxl import Workbook


# класс, который позволит считать периметр и площадь


class MyClass:
    def __init__(self, a, b, name='квадрат'):
        self.name = name
        self.a = int(a)
        if isinstance(b, str):
            b = int(b)
        self.b = b

    def get_perimeter(self):  # метод внутри класса, который работает с внутренними параметрами
        return self.a + self.b

    @staticmethod
    def get_perimeter_static(a, b):  # метод внутри класса, который работает с внутренними параметрами
        return a + b

    def get_square(self):  # метод внутри класса, который работает с внутренними параметрами
        return self.a * self.b

    @staticmethod
    def get_square_static(a, b):  # метод внутри класса, который работает с внутренними параметрами
        return a * b


# oop Калькулятор

class MyCalculator:
    def __init__(self, val1: float, val2: float):
        self.val1 = val1
        self.val2 = val2

    def sum1(self):
        return self.val1 + self.val2

    @staticmethod
    def sum1_static(val1: float, val2: float):
        print(val1 + val2)


if __name__ == "__main__":
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = '12'
    workbook.save('new.xlsx')

    print(workbook)
    print(type(workbook))

    print(MyClass)
    myclass = MyClass(name="Прямоугольник", b=15, a=15)  # создание экземпляра класса
    print(myclass)
    print(type(myclass))
    print(myclass.a)
    print(myclass.b)
    print(myclass.name)
    print(myclass.get_perimeter())
    print(myclass.get_square())
    print(MyClass.get_perimeter_static(17, 18))

    MyCalculator.sum1_static(20, 30.0)
    value = MyCalculator(30, 21.0).sum1()
    print(value)
else:
    pass


#########################################################################################################

class Mother:
    val1 = 12

    def __init__(self, val1, name="Мама") -> None:
        self.name = name
        self.val = val1
        self._val = val1  # защищённый
        self.__val2 = val1 + 5  # приватный

    def get_value2(self):
        return 10

    def __str__(self):
        return self.name


class Father:
    val2 = 13

    def __init__(self, val1) -> None:
        self.val1 = val1

    def get_value(self):
        return 8

    def __str__(self):
        return self.val1


class Child(Mother, Father):

    def __init__(self, val1) -> None:
        super().__init__(val1)

    def get_value1(self):
        return 5


print(Mother.val1)
a1 = Mother(10)
print(a1.val)
# print(a1._Mother__val2)

ch1 = Child(10)
print(ch1.get_value())

########################################################################################################################


# class Base(object):
#     def __init__(self):
#         print
#         "Base created"
#
#
# class ChildA(Base):
#     def __init__(self):
#         Base.__init__(self)
#
#
# class ChildB(Base):
#     def __init__(self):
#         super(ChildB, self).__init__()
#
#
# ChildA()
# ChildB()

########################################################################################################################

import math
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

filename = 'users.xlsx'
filedir = 'home_work'

workbook = openpyxl.load_workbook(filedir + "/" + filename)
worksheet = workbook.active

max_row = worksheet.max_row
max_column = worksheet.max_column

workers = []


class Worker:  # определение (создание базовых параметров) класса
    def __init__(self, first_name: str, second_name: str, patronymic: str, position,
                 category="Рабочий"):  # магический метод для инициализации (создания) класса
        self.value_first_name = first_name  # свойство (переменная) внутри класса - изменяемая
        self.value_second_name = second_name  # свойство (переменная) внутри класса - изменяемая
        self.value_patronymic = patronymic  # свойство (переменная) внутри класса - изменяемая
        self.value_position = position  # свойство (переменная) внутри класса - изменяемая
        self.value_category = category  # свойство (переменная) внутри класса - изменяемая

        self.radius = 0

    def print_full_name(self):
        print(f'{self.value_second_name} {self.value_first_name}')

    def get_full_name(self):
        return f'{self.value_second_name} {self.value_first_name}'

    def calcuate(self, side=14):
        self.radius = side ** 2


for row in range(1, max_row + 1):
    worker = []

    for column in range(1, 5 + 1):
        obj = worksheet.cell(row=row, column=column)  # метод (функция в классе) который получает объект по координатам
        # print(obj)
        # print(type(obj))

        value = obj.value  # свойство (переменная в классе) которое получает значение ячейки

        if not value:
            value = ''
        worker.append(value)

    # print(worker)
    # print('читаю нового работника!')
    worker_obj = Worker(
        first_name=worker[1],
        second_name=worker[0],
        patronymic=worker[2],
        position=worker[3],
        category=worker[4],
    )

    # worker_obj.print_full_name()
    full_name = worker_obj.get_full_name()
    # print(full_name)

    workers.append(worker_obj)
    # workers.append(worker)
# print(workers)

print(workers[1].value_patronymic)

print(f'{workers[1].value_second_name} {workers[1].value_first_name}')  # Исагулов  Куаныш
workers[1].print_full_name()  # Исагулов  Куаныш
print(workers[1].get_full_name())  # Исагулов  Куаныш

external_value_variable = 1

workbook = Workbook()
print(workbook)
print(type(workbook))


class CustomWorkerClass1:  # определение (создание базовых параметров) класса
    def __init__(self):  # магический метод для инициализации (создания) класса
        self.VALUE_CONSTANT = 321  # свойство (переменная) внутри класса - константа (не изменяется)


print(CustomWorkerClass1)
print(type(CustomWorkerClass1))

worker = CustomWorkerClass1()  # создание экземпляра класса CustomWorkerClass
print(worker)
print(type(worker))
print(worker.VALUE_CONSTANT)  # получение свойства данного экземпляра данного класса


class CustomWorkerClass2:  # определение (создание базовых параметров) класса
    def __init__(self, external_value_variable2=15):  # магический метод для инициализации (создания) класса
        self.VALUE_CONSTANT = 321  # свойство (переменная) внутри класса - константа (не изменяется)
        self.value_variable = external_value_variable2  # свойство (переменная) внутри класса - изменяемая


worker = CustomWorkerClass2(external_value_variable2=156)  # создание экземпляра класса CustomWorkerClass
print(worker)
print(type(worker))
print(worker.VALUE_CONSTANT)  # получение свойства данного экземпляра данного класса
print(worker.value_variable)  # получение свойства данного экземпляра данного класса

print("\n\n\n**********\n\n\n")


# класс для расчёта периметров и площадей
class MyClass:
    def __init__(self, name: str, side1: float, side2: float, side3: float, side4=0.0):
        self.name = name
        self.side1 = side1
        self.side2 = side2
        self.side3 = side3
        self.side4 = side4

        self.radius = 0

    def print_name(self):
        print(self.name)

    def get_perimeter(self):
        perimeter_value = self.side1 + self.side2 + self.side3 + self.side4
        return perimeter_value

    def get_square_with_multiply(self, multiply=1.0):
        if self.side4 > 0:
            # фигуры с 4 сторонами
            return (self.side1 * self.side2) * multiply
        else:
            side1 = self.side1
            print(f'side1: {side1}')
            side2 = self.side2
            print(f'side2: {side2}')
            side3 = self.side3
            print(f'side3: {side3}')
            p = (side1 + side2 + side3) / 2
            print(p)
            value1 = p * (p - side1) * (p - side2) * (p - side3)
            print(value1)
            s = math.sqrt(value1)
            print(s)
            return s * multiply

    def calcuate(self, side=14):
        self.radius = side ** 2


figure1 = MyClass(name="квадрат", side4=10, side2=10, side3=10, side1=10)  # квадрат
figure1.print_name()

perimeter_perimeter_figure1 = figure1.get_perimeter()
print(perimeter_perimeter_figure1)

square_perimeter_figure1 = figure1.get_square_with_multiply(multiply=2.5)
print(square_perimeter_figure1)

figure2 = MyClass(name="прямоугольник", side4=20, side2=20, side3=7, side1=7)  # прямоугольник
figure2.print_name()

perimeter_perimeter_figure2 = figure2.get_perimeter()
print(perimeter_perimeter_figure2)

square_perimeter_figure2 = figure2.get_square_with_multiply(0.75)
print(square_perimeter_figure2)

figure3 = MyClass(name="треугольник", side2=20, side3=15, side1=7)  # треугольник
figure3.print_name()

perimeter_perimeter_figure3 = figure3.get_perimeter()
print(perimeter_perimeter_figure3)

square_perimeter_figure3 = figure3.get_square_with_multiply(0.75)
print(square_perimeter_figure3)

# new_data = float(input("Введите любое число: "))  # возвращает данные, которые ввёл пользователь в виде строки
# print(new_data)
# print(type(new_data))

# side1 = float(input("Введите первую сторону: "))

# user_string = input("Введите через запятую стороны объекта (12, 35, 65...): ")  # '12, 56,89, 40'
# print(user_string)
# print(type(user_string))
# # user_string = '12, 56,89, 40'
#
# # обработка ошибок
#
# sides = []
# for x in user_string.split(sep=','):
#     value = float(str(x).strip())
#     sides.append(value)
# print(sides)
#
# # sides = [x for x in]
# if len(sides) == 3:
#     sides.append(0.0)
# elif len(sides) < 3:
#     print(f"Вы ввели только {len(sides)} сторону(-ы)!")
# print(sides)
#
# try:
#     figure4 = MyClass(name="Новый объект", side1=sides[0], side2=sides[1], side3=sides[2], side4=sides[3])  # Новый объект
# except:
#     figure4 = MyClass(name="Новый объект", side1=sides[0], side2=sides[1], side3=sides[2])  # Новый объект
#
#
# figure4.print_name()
#
# perimeter_perimeter_figure4 = figure4.get_perimeter()
# print(perimeter_perimeter_figure4)
#
# square_perimeter_figure4 = figure4.get_square_with_multiply(0.5)
# print(square_perimeter_figure4)


print("\n\n\n**********\n\n\n")


class MyCalculator:
    def __init__(self, val1: float, val2: float):
        try:
            self.val1 = float(val1)
        except Exception as error:
            self.val1 = input('Введите первое значение ещё раз: ')
        self.val2 = val2

    def summ2(self):
        try:
            return float(self.val1) + float(self.val2)
        except Exception as error:
            print(error)
            return 0.0

    def multiply(self):
        return float(self.val1) * float(self.val2)

    @staticmethod
    def multiply_static(val1, val2):
        return float(val1) * float(val2)

    @staticmethod  # декоратор, который делает метод в классе статическим(без параметра селф и инициализации)
    def summ(val1: float, val2: float):  # статический метод
        return val1 + val2


# инициализация калькулятора
# summ1 = MyCalculator('sdfsd1sdg2', "16")
# print(summ1.summ2())
# print(summ1.multiply())
# print(MyCalculator.summ(15, 17))
# print(MyCalculator.multiply_static(15, 17))


# def calc_3(number1, number2, operation="-"):
#     # print(number1, number2, operation)
#     if operation == "+":
#         return number1 + number2
#     if operation == "-":
#         return number1 - number2
#     if operation == "*":
#         return number1 * number2
#     if operation == "/":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 / number2
#     if operation == "**":
#         return number1 ** number2
#     if operation == "//":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 // number2
#     if operation == "sqrt":
#         return math.sqrt(number1)
#     if operation == "%":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 % number2

class MyTree:
    def __init__(self, name):
        self.age = 1
        self.name = name

    def drow(self, value_: float):  # рост дерева(старение)
        self.age += value_

    @staticmethod
    def drow_static(value_: float, name: str):
        age = 1
        age += value_
        # tree = {"name": name, "age": age}
        # return tree
        return dict(name=name, age=age)

tree1 = MyTree("дерево 1")  # 1
print('возраст ', tree1.name, ' = ', tree1.age)  # 1
tree1.drow(2)  # 3
print('возраст ', tree1.name, ' = ', tree1.age)  # 3
tree1.drow(2)  # 5
print('возраст ', tree1.name, ' = ', tree1.age)  # 5

tree2 = MyTree("дерево 2")  # 1
print('возраст ', tree2.name, ' = ', tree2.age)  # 1
tree2.drow(1)  # 2
print('возраст ', tree2.name, ' = ', tree2.age)  # 2

tree_dict = MyTree.drow_static(2, "дерево 3")
print(f"имя дерева: '{tree_dict['name']}', возраст дерева: '{tree_dict['age']}'")
# имя дерева: 'дерево 3', возраст дерева: '3'

########################################################################################################################

# class Meter():
#     def __init__(self, dev):
#         self.dev = dev
#     def __enter__(self):
#         #ttysetattr etc goes here before opening and returning the file object
#         self.fd = open(self.dev, MODE)
#         return self
#     def __exit__(self, type, value, traceback):
#         #Exception handling here
#         close(self.fd)
#
# meter = Meter('dev/tty0')
# with meter as m:
#     #here you work with the file object.
#     m.fd.read()
import datetime


# Процедурное программирование               Паскаль
# Функциональное программирование            F#
# Объектно-ориентированное программирование  C#, Делфи, Python, c++


#                     объект - видимость, рендеринг, коллизии (физики)
#                     техника - скорость, масса, может с ними взаимодействовать
#         сухопутные             водоплавающие  - точки, где можно перемещаться
#  Велосипед Машина Мотоцикл  гидроскутер  Лодка     - текстуры, цвет


# print(type(10))
# print(type(True))
# print(type([True, 12]))

class Obj:
    is_start_rendering1 = True
    visible = False  # публичное свойство экземпляра класса
    _visible = True  # защищённое свойство экземпляра класса
    __visible = True  # приватное свойство экземпляра класса

    def __init__(self, mass=10.0, is_start_rendering=True):  # магический метод(функция)
        self.mass = mass
        self.is_start_rendering1 = True
        self.is_start_rendering = is_start_rendering
        self.mass2 = 12
        self.__mass3 = 12

    def get_visibility(self) -> bool:  # метод
        """
        Return state of is visibile obj
        :return: bool
        """
        return self._visible

    def set_visibility(self, is_checked: bool) -> None:
        self._visible = is_checked
        print(self.mass2)

    def get_mass(self):
        return self.mass

    # def switch_visibility(self) -> None:  # метод
    #     """
    #     Этот метод должен переключать видимость
    #     :return: None
    #     """
    #     self._visible = not self._visible

    # def switch_and_return_visibility(self) -> bool:  # метод
    #     """
    #     Этот метод должен переключать видимость
    #     :return: None
    #     """
    #     self._visible = not self._visible
    #     return self._visible

    # def print_visibility(self) -> None:
    #     print(self._visible)


class Vehicle(Obj):
    def __init__(self, speed: float):
        super().__init__(mass=100)
        self.speed = speed
        self.mass = 200

    def get_parent_mass(self):
        return super().get_mass()

    # как получить значение такого атрибута, но от родителя
    # использование  public || private || protected
    # множественное наследование и "проблема ромба"
    # бинарное дерево


class Some:
    pass


som1 = Some()

veh1 = Vehicle(speed=50.2)
print(veh1.mass)
print(veh1.get_parent_mass())


#
# print(Obj.visible)
# print(Obj._visible)
# # print(Obj.__visible)
#
# new_obj_1 = Obj()  # создание экземпляра
# print(new_obj_1)
# print(new_obj_1.visible)
# print(new_obj_1._visible)
# # print(new_obj_1.__visible)
# print(type(new_obj_1))

# new_obj_1 = Obj()
#
# print(new_obj_1.get_visibility())
#
# # new_obj_1.switch_visibility()
#
# print(new_obj_1.get_visibility())
#
# new_obj_1.set_visibility(is_checked=not new_obj_1.get_visibility())
#
# print(new_obj_1.get_visibility())


class Parall:
    def __init__(self, side1: int, side2: int):
        self.side1 = side1
        self.side2 = side2
        self.side3 = side1 / 3 * side2

    def check_is_square(self) -> bool:
        # if self.side1 == self.side2:
        #     return True
        # else:
        #     return False
        return self.side1 == self.side2

    def get_perimeter(self) -> int:
        return (self.side1 + self.side2) * 2

    def get_square(self) -> int:
        if self.check_is_square():
            return self.side1 ** 2  # квадрат
        else:
            return self.side1 * self.side2  # многоугольник

    def __add__(self, other) -> int:  # +
        if isinstance(other, Parall):
            return self.get_perimeter() + other.get_perimeter()
        else:
            raise TypeError

    def sum_self(self, other) -> int:  # +
        if isinstance(other, Parall):
            return self.get_perimeter() * other.get_perimeter()
        else:
            raise TypeError


parall1 = Parall(side1=2, side2=3)
parall2 = Parall(side1=30, side2=30)
parall3 = Parall(side1=50, side2=30)


# print(parall1 + 10)

# list1 = [Parall(side1=x + 1, side2=x + 2) for x in range(1, 101)]
# sum1 = 0
# print(list1)
# for i in list1:
#     sum1 += i.get_square()
# print(sum1)
# print(parall1.get_perimeter())

# if parall1.check_is_square():
#     print('Это квадрат!')
# else:
#     print('Это прямоугольник!')
#
# if parall2.check_is_square():
#     print('Это квадрат!')
# else:
#     print('Это прямоугольник!')


class HelpPython:
    class Time:
        @staticmethod
        def get_current_time(timezone: int, formatting="23:59:59"):
            import datetime
            datetime = datetime.datetime.now()

            match formatting:
                case "23:59:59":
                    return datetime.strftime('%H:%M:%S')
                case "23:59:59.999":
                    return datetime.strftime('%H:%M:%S ...')
                case "23:59":
                    return datetime.strftime('%H:%M')
                case _:
                    return datetime

        @staticmethod
        def get_different_times_in_seconds(datetime1: datetime.datetime, datetime2: datetime.datetime) -> int:
            # datetime1 - datetime2
            return 0

    class Variable:
        class Types:
            @staticmethod
            def example():
                value1 = 10  # - 100
                value2 = 10.0  # -10.6
                value3 = True  # False
                value4 = ""
                value5 = "10"
                value6 = "Hello"
                str1 = b'\x01\x02\x03\x04\x05'
                str2 = "\n \t  I'am "
                str3 = """\n \t  I'am """
                str4 = '''\n \t  I'am '''
                str5 = '\n \t \\ I\'am '
                str6 = r"\n\tC:\Users\bogdan\Desktop\Django\Курс лекций Django1\Python"

        @staticmethod
        def convert_to_string(value) -> str:
            # return str(value)
            # return str(value).format(1)
            return f"{value}"

        @staticmethod
        def get_type_variable(var) -> type:
            return type(var)

        @staticmethod
        def get_value_from_dict(dictionary: dict, value: str, default_value, is_logging=False):
            try:
                # return dictionary.get("username", None)  # не вызывает исключения при отсутвии ключа
                return dictionary[value]
            except Exception as error:
                if is_logging:
                    print(f"Ошибка: {error}")
                return default_value

    class Arifmetic:
        @staticmethod
        def example():
            import math  # много математики

            val1 = -10
            val2 = 15

            result = math.sqrt(16)  # int(math.sqrt(16))
            print(type(result > 3.9))

            print(result > 3.9)

            abs(val1)

            print(10 * 4)
            print(10 ** 4)
            print(10 / 5)
            print(10 // 5)  # int(10 / 5) round(10 / 3, 2) 3.33

            print(16 ** 0.5)

            print(7 % 2)  # 1
            print(6 % 2 == 0)  # чётное

    class ExceptionHelps:
        class MyException1(Exception):
            def __init__(self, exception_text: str, sector: str, level_error: int, critical: bool):
                self.exception_text = exception_text
                self.datetime = datetime.datetime.now()
                self.sector = sector
                self.level_error = level_error
                self.critical = critical

            def print_error(self):
                print(f"{self.sector}: {self.exception_text} {self.datetime}")


dict1 = {"name": "Python"}
print(HelpPython.Variable.get_value_from_dict(dictionary=dict1, value="name", default_value="C++", is_logging=False))


class MyException(Exception):
    def __init__(self, exception_text: str, sector: str, level_error: int, critical: bool):
        self.exception_text = exception_text
        self.datetime = datetime.datetime.now()
        self.sector = sector
        self.level_error = level_error
        self.critical = critical

    def print_error(self):
        print(f"{self.sector}: {self.exception_text} {self.datetime}")


def test1(a, b):
    try:
        if b == 0:
            raise HelpPython.ExceptionHelps.MyException1(exception_text="кто-то накосячил с кодом",
                                                         sector="test1 Вычислительное ядро",
                                                         level_error=3, critical=False)
        c = a / b
        return c
    except HelpPython.ExceptionHelps.MyException1 as error:
        print(f"MY {error}")
        print(error.sector)
        print(error.print_error())
        return a
    except Exception as error:
        print(f"123124124 {error}")
        return None


print(test1(9, 0))


class BinaryTree:
    def __init__(self, data: int):
        self.left = None
        self.right = None
        self.root = data

    def get_data(self):
        return self.root

    def insert_new_data(self, data: int):
        if self.root:
            if data < self.root:

                # левый
                if self.left is not None:
                    self.left.insert_new_data(data)
                else:
                    self.left = BinaryTree(data)
                # левый

            elif data > self.root:

                # правый
                if self.right is not None:
                    self.right.insert_new_data(data)
                else:
                    self.right = BinaryTree(data)
                # правый

            else:
                print('Значение повторяется')

        else:
            self.root = data

    def print_all_edges(self):
        print(self.root)
        print(self.left)
        print(self.right)


tree1 = BinaryTree(1)  # создание экземляра класса - создание объекта
tree1.insert_new_data(2)
tree1.insert_new_data(3)
tree1.insert_new_data(9)
tree1.insert_new_data(6)
tree1.insert_new_data(11)
tree1.insert_new_data(15)
print(f"root: {tree1.root}")
print(f"root: {tree1.right.root}")

print(f"root: {tree1.right.right.root}")

print(f"root: {tree1.right.right.right.root}")

print(f"root: {tree1.right.right.right.right.root}")
print(f"root: {tree1.right.right.right.left.root}")

print(f"root: {tree1.right.right.right.right.right.root}")

# class BinaryTree2:
#     def __init__(self, data: int):
#         self.left = None
#         self.right = None
#         self.root = data
#
#
# tree1 = BinaryTree2(1)  # создание экземляра класса - создание объекта
# tree2 = BinaryTree2(2)  # присвоение в левую ветку нового объекта
# tree1.left = tree2  # присвоение в правую ветку нового объекта
# tree1.right = BinaryTree2(3)  # присвоение в правую ветку нового объекта
# tree1.right.right = BinaryTree2(6)
# print(tree1)
# print(f"root: {tree1.root}")
# print(f"root: {tree1.left.root}")

index = 0
def substr(val: int):
    global index
    index += 1
    print(f'вход номер {index}, значение {val}')

    if val == 1:
        return val
    else:
        return substr(val//2)

print(substr(12))

# CreateReadUpdateDelete
# CRUD

# Date and time is: 2021-07-03 16:21:12.357246
# Timestamp is: 1625309472.357246

########################################################################################################################

import openpyxl
from openpyxl import Workbook


class GrandMother:
    eyes_color = 'brown'


class Father:
    height = 180
    eyes_color = 'gray'


class Mother(GrandMother):
    __eyes_color = 'blue'

    def __init__(self, hair_color: str, multiply_value=2.0):
        self.multiply_value = multiply_value
        self.hair_color = hair_color

    def give_money(self, value: float):
        return value * self.multiply_value


class Child(Mother, Father):
    # eyes_color = 'green'

    def __init__(self, hair_color: str, multiply_value):
        super().__init__(hair_color, multiply_value)

    def give_money(self, value: float):
        return value * self.multiply_value * 1000


m1 = Mother('red', multiply_value=3)
print(m1.give_money(10))

print(isinstance(m1, Child))

if type(m1) == GrandMother:
    print('ok')
else:
    print('not ok')

# print(m1.eyes_color)


child = Child('red', multiply_value=0.5)
print(f'Child: {child._Mother__eyes_color}')
print(f'Child: {child.hair_color}')

# print(child.give_money(20))

print(child.height)

wb = Workbook()


########################################################################################################################

import openpyxl


class MyExcel:
    def __init__(self, filename=None):
        if filename is None:
            self.workbook = MyExcel.create_new_workbook()
        else:
            self.workbook = MyExcel.load_workbook(filename=filename)
        self.worksheet = self.workbook.active
        self.name = filename

    def read_all_data(self, sheetname = '') -> list[list]:
        # TODO тут нужно дописать чтение определенного листа
        arr = []
        for row in range(1, self.worksheet.max_row + 1):
            arr_tmp = []
            for col in range(1, self.worksheet.max_column + 1):
                val = self.worksheet.cell(row=row, column=col).value
                if val is None:
                    val = ""
                arr_tmp.append(val)
            arr.append(arr_tmp)
        return arr

    # def __str__(self):
    #     return self.read_all_data()

    def fill_sheet_from_matrix(self, arr: list[list]) -> None:
        for i in range(1, len(arr) + 1):
            for j in range(1, len(arr[i-1]) + 1):
                self.worksheet.cell(row=i, column=j, value=arr[i-1][j-1])

    def save(self, name=None) -> None:
        if name is None:
            self.workbook.save(self.name)
        else:
            self.workbook.save(name)

    @staticmethod
    def load_workbook(filename: str) -> openpyxl.Workbook:
        return openpyxl.load_workbook(filename)

    @staticmethod
    def create_new_workbook() -> openpyxl.Workbook:
        return openpyxl.Workbook()


wb1 = MyExcel(filename='data.xlsx')
print(wb1)

arr_tmp = [[0, 4, 5], [3, 2, 6]]
wb1.fill_sheet_from_matrix(arr_tmp)
print(wb1)
wb1.save(name='data_new.xlsx')

########################################################################################################################
