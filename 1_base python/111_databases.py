import psycopg2

try:
    connection = psycopg2.connect(
        user="postgres",
        password="31284bogdan",
        host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
        port="5432",
        dbname="example",
    )
    cursor = connection.cursor()
    query_string = '''
DELETE FROM public.example_table
WHERE id = 888
    '''
    cursor.execute(query_string)
    connection.commit()
except Exception as error:
    print(error)
finally:
    cursor.close()
    connection.close()

import psycopg2
from api import utils

try:
    connection = psycopg2.connect(
        user="postgres",
        password="31284bogdan",
        host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
        port="5432",
        dbname="example",
    )
    cursor = connection.cursor()
    query_string = '''
SELECT * FROM public.example_table
WHERE age = 50
ORDER BY id ASC
    '''
    cursor.execute(query_string)
    records = cursor.fetchall()
    print(records)
    print(type(records))
    for i in records:
        print(i)
        print(type(i))
except Exception as error:
    print(error)
finally:
    cursor.close()
    connection.close()




# obj = utils.SQL(
#     autocommit=True,
#     user="postgres",
#     password="31284bogdan",
#     host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
#     port="5432",
#     dbname="example",
# )
# records = obj.execute('''
# SELECT * FROM public.example_table
# WHERE age = 50
# ORDER BY id ASC
#     ''')
# print(records)
#
#
# records = obj.execute('''
# SELECT * FROM public.example_table
# WHERE id > 50
# ORDER BY id ASC
#     ''')
# print(records)

import psycopg2

try:
    connection = psycopg2.connect(
        user="postgres",
        password="31284bogdan",
        host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
        port="5432",
        dbname="example",
    )
    connection.autocommit = True
    cursor = connection.cursor()
    query_string = '''
UPDATE public.example_table
SET credits = '0.0'
WHERE married = 'true'
    '''
#     query_string = '''
# CREATE TABLE public.table1
# (
#     title text NOT NULL,
#     id serial NOT NULL,
#     PRIMARY KEY (id)
# );
#
# ALTER TABLE IF EXISTS public.table1
#     OWNER to postgres;
#     '''
    cursor.execute(query_string)
    # connection.commit()
except Exception as error:
    print(error)
finally:
    cursor.close()
    connection.close()

import psycopg2

try:
    # number = input("Введите номер id: ")

    class RowObj:
        def __init__(self, row: tuple):
            self.username = row[0]
            self.age = row[1]
            self.married = row[2]
            self.credits = row[3]
            self.id = row[4]

    arr = [
        ('x', 12, 'false', 888.888, 85),
        ('y', 13, 'false', 888.888, 86),
        ('z', 14, 'false', 888.888, 87),
        ('a', 15, 'false', 888.888, 88),
        ('b', 16, 'false', 888.888, 89),
    ]
    connection = psycopg2.connect(
        user="postgres",
        password="31284bogdan",
        host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
        port="5432",
        dbname="example",
    )
    connection.autocommit = True
    cursor = connection.cursor()
    for row in arr:
        obj = RowObj(row=row)
        query_string = f'''INSERT INTO public.example_table 
        (username, age, married, credits, id) values 
        ('{obj.username}', {obj.age}, {obj.married}, {obj.credits}, {obj.id})'''
        cursor.execute(query_string)
        # connection.commit()
except Exception as error:
    print(error)
finally:
    cursor.close()
    connection.close()

########################################################################################################################

import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter

# Read all data from SQL
conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()
cur.execute("""
SELECT * FROM public.products
ORDER BY id ASC
""")
records = cur.fetchall()
conn.close()

# Write all data to .xlsx
book = openpyxl.Workbook()
sheet = book.active

titles = ["Заголовок 1", "Заголовок 2", "Заголовок 3"]
index_col = 1
for title in titles:
    sheet[f"{get_column_letter(index_col)}1"] = str(title)
    index_col += 1

row_index = 2
for row in records:
    col_index = 1
    for column in row:
        sheet[f"{get_column_letter(col_index)}{row_index}"] = str(column)
        col_index += 1
    row_index += 1
book.save('excel_data/new_table.xlsx')

########################################################################################################################

import psycopg2
import openpyxl

# Read all data from .xlsx
book = openpyxl.load_workbook("excel_data/tab.xlsx")
sheet = book.active
global_rows = []
for row in range(2, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)


# Example
# conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
# cur = conn.cursor()
# query_string = f"""
# INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
# VALUES ('tovar', 'gruppa', 'postavshick', '2022-06-06', 'region', '500', '600', 'true')
# """
# cur.execute(query_string)
# conn.commit()


# Insert all data to SQL
class Record:
    def __init__(self, row: list):
        self.tovar = str(row[0])
        self.gruppa = str(row[1])
        self.postavshick = str(row[2])
        date_post = str(row[3]).split(" ")[0]  # '2010-12-4'
        year = date_post.split('-')[0]
        month = int(date_post.split('-')[1])  # '09' -> 9
        if int(month) < 10:
            month = f"0{month}"
        day = int(date_post.split('-')[2])
        if int(day) < 10:
            day = f"0{day}"
        self.date_post = str(f"{year}-{month}-{day}")
        self.region = str(row[4])
        try:
            self.prodashi = int(row[5])
        except Exception as error:
            print(error)
            self.prodashi = 0
        try:
            self.sbit = int(row[6])
        except Exception as error:
            print(error)
            self.sbit = 0
        pribil = str(row[7])
        if pribil.lower() == "да":
            self.pribil = True
        else:
            self.pribil = False


conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()
for row in global_rows:
    obj = Record(row=row)
    try:
        query_string = f"""
        INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
        VALUES ('{obj.tovar}', '{obj.gruppa}', '{obj.postavshick}', '{obj.date_post}', '{obj.region}', '{obj.prodashi}', 
        '{obj.sbit}', '{obj.pribil}')
        """
        cur.execute(query_string)
        conn.commit()
    except Exception as error:
        print(error)
        print(row)
conn.close()

########################################################################################################################

import psycopg2

# CRUD

# Read - чтение(сортировка и фильтрация) из базы и запись в эксель файл
# Create - вставка новых данных (из эксель файл)
# Delete - удаление строк по условиям
# Update - обновление из базы и запись в эксель файл

# READ (SELECT)
############################################################################

# Connect to your postgres DB
conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")  # localhost (127.0.0.1 / 192.168.1.121)

# Open a cursor to perform database operations
cur = conn.cursor()

# Execute a query

# read data
cur.execute("""
SELECT * FROM public.example_table
WHERE age > 19
ORDER BY Age ASC, credits DESC
""")

# Retrieve query results
records = cur.fetchall()

print(records)
print(type(records))

for i in records:
    print(i)
    print(type(i))

conn.close()

# CREATE (INSERT)
############################################################################
conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()

# ('t', 25, True, 30000.6, 1)

new_arr = [
    ['w', 25, 'true', 30000.6, 1],
    ['b', 50, 'false', 30.6, 1],
    ['y', 25, 'true', 30000.6, 1],
    ['r', 75, 'false', 500.6, 0],
    ['u', 88, 'true', 30000.6, 1],
    ['3', 25, 'true', 305.6, 0],
]

# create data
index = 12
for i in new_arr:
    query_string = f"""
    INSERT INTO public.example_table (username, age, married, credits, id) 
    VALUES ('{i[0]}', {i[1]}, {i[2]}, {i[3]}, {index})
    """
    index += 1

    cur.execute(query_string)
    conn.commit()  # применение данных после изменений

conn.close()

# DELETE (DELETE)
############################################################################

conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()

query_string = """
    DELETE FROM public.example_table
    WHERE age <= 50 and married = 'true';
    """

#

cur.execute(query_string)
conn.commit()
conn.close()

# UPDATE (UPDATE)
############################################################################

conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()

query_string = """
UPDATE public.example_table
SET credits = '666.66' 
WHERE id = 2;
"""

#

cur.execute(query_string)
conn.commit()
conn.close()

# CREATE TABLE
############################################################################

conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()

query_string = """
CREATE TABLE public.products
(
    tovar text,
    gruppa text,
    postavshick text,
    date_post date,
    region text,
    prodashi integer DEFAULT 0,
    sbit integer DEFAULT 0,
    pribil boolean NOT NULL DEFAULT false,
    id SERIAL,
    PRIMARY KEY (id)
);
ALTER TABLE IF EXISTS public.products
    OWNER to postgres;
"""
cur.execute(query_string)
conn.commit()
conn.close()

########################################################################################################################


