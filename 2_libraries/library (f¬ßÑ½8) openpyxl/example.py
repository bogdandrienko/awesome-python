import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, GradientFill, Alignment, Side

# workbook = Workbook()  # для создания в памяти экземпляра Workbook
# worksheet = workbook.active

workbook = openpyxl.load_workbook('users.xlsx')
worksheet = workbook.active

#            0    1
new_list = [123, 124]
print(new_list[0])

new_dict = {"first": 123, 23: 666, "23": 666}
print(new_dict["first"])

value = worksheet["A56"].value
print(value)
print(type(value))

new_set = {12, 135, 656, 12}
print(new_set)
print(type(new_set))

new_tuple = (12,)

workers = []
for row in range(1, 2150):
    coordinates = "A" + str(row)  # 1 способ сложения строк
    # print(coordinates)
    coordinates = f"A{row}"  # 2 способ сложения строк
    # print(coordinates)
    coordinates = "A{}".format(row)  # 3 способ сложения строк
    # print(coordinates)
    col = "A"
    coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
    # print(coordinates)
    # txt1 = "My name is {fname}, I'm {age}".format(fname="John", age=36)
    # txt2 = "My name is {0}, I'm {1}".format("John", 36)
    # txt3 = "My name is {}, I'm {}".format("John", 36)

    worker = []
    for col in "ABCDEF":
        coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
        value = worksheet[coordinates].value
        if value is not None:
            # print(value)
            worker.append(value)
            pass
    workers.append(worker)


# print(workers)


class Worker:
    def __init__(self, first_name: str, second_name, patronymic, id, position, category):
        self.first_name = first_name
        self.second_name = second_name
        self.patronymic = patronymic
        self.id = id
        self.position = position
        self.category = category

    def get_full_name(self):
        return f"{self.first_name} {self.second_name}"


max_row = worksheet.max_row + 1
max_column = worksheet.max_column + 1

workers = []
for row in range(2, max_row):
    worker = []
    for column in range(1, max_column):
        value = worksheet.cell(row=row, column=column).value
        worker.append(value)
    # print(worker)
    # print(type(worker))
    worker = Worker(
        first_name=worker[0],
        second_name=worker[1],
        patronymic=worker[2],
        id=worker[3],
        position=worker[4],
        category=worker[5]
    )
    workers.append(worker)
    # print(worker)
    # print(type(worker))

print(workers[2])

workbook = Workbook()  # для создания в памяти экземпляра Workbook
worksheet = workbook.active

worksheet[f"A1"] = "Фамилия Имя"
worksheet.merge_cells('B2:C6')
a1 = worksheet[f"A1"]
a1.font = Font(color="00FF6600", bold=True, sz=28)

index = 1
for worker in workers:
    index += 1
    worksheet[f"A{index}"] = worker.get_full_name()

workbook.save('new1.xlsx')

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:F4')
top_left_cell = ws['B2']
top_left_cell.value = "My Cell"
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
top_left_cell.font  = Font(b=True, color="FF0000")
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
wb.save("styled.xlsx")

###############################################################################

class ExcelClass:
    @staticmethod
    def workbook_create():
        workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def workbook_load(excel_file: str):
        workbook = openpyxl.load_workbook(excel_file)
        return workbook

    @staticmethod
    def workbook_activate(workbook):
        sheet = workbook.active
        return sheet

    @staticmethod
    def workbook_save(workbook, excel_file: str):
        try:
            os.remove(excel_file)
        except Exception as error:
            pass
        try:
            workbook.save(excel_file)
        except Exception as error:
            print(f'\n ! Please, close the excel_file! \n: {excel_file} | {error}')

    @staticmethod
    def workbook_close(workbook):
        openpyxl.Workbook.close(workbook)

    @staticmethod
    def set_sheet_title(sheet, page_name='page 1'):
        sheet.title = page_name

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        value = str(sheet[str(col).upper() + str(row)].value)
        if value == 'None' or value is None:
            value = ''
        else:
            value = str(value)
        return value

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        if value == 'None' or value is None:
            value = ''
        sheet[str(col) + str(row)] = str(value)

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)

    @staticmethod
    def get_max_num_rows(sheet):
        return int(sheet.max_row)

    class Example:
        @staticmethod
        def example_read_from_excel_file_col_int():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            max_num_cols = 10
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in range(1, max_num_cols + 1):
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_read_from_excel_file_col_char():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            char_cols = 'ACDF'
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in char_cols:
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_write_to_excel_file():
            global_list = [
                ['title_1', 'title_2', 'title_3'],
                ['body_1_1', 'body_1_2', 'body_1_3'],
                ['body_2_1', 'body_2_2', 'body_2_3'],
                ['body_3_1', 'body_3_3', 'body_3_3'],
            ]

            excel_file = 'import.xlsx'
            workbook = ExcelClass.workbook_create()
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_list:
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=row.index(value) + 1,
                        row=global_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=excel_file)
            
#############################################################################################


class ExampleClass:
    @staticmethod
    def example():
        # Сравнение данных из 1с и данных в базе данных енбека
        def extra_enbek():
            start_time = time.time()
            print('start')

            input_1 = '1c.xlsx'
            input_2 = 'enbek.xlsx'
            output_1 = 'final.xlsx'

            workbook = ExcelClass.workbook_load(excel_file=input_1)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

            global_workers_id_list = []
            for row in range(1, max_num_rows + 1):
                global_workers_id_list.append(
                    ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(11), row=row, sheet=sheet)
                )
            print(global_workers_id_list)
            ExcelClass.workbook_close(workbook=workbook)

            workbook = ExcelClass.workbook_load(excel_file=input_2)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            global_workers_enbek_list = []
            for row in range(1, max_num_rows + 1):
                local_workers_enbek_list = []
                for col in range(1, 21 + 1):
                    local_workers_enbek_list.append(
                        ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(col), row=row, sheet=sheet)
                    )
                if local_workers_enbek_list[1] in global_workers_id_list:
                    local_workers_enbek_list.append('+')
                else:
                    local_workers_enbek_list.append('-')
                global_workers_enbek_list.append(local_workers_enbek_list)
            print(global_workers_enbek_list)
            ExcelClass.workbook_close(workbook=workbook)

            workbook = ExcelClass.workbook_load(excel_file=output_1)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_workers_enbek_list:
                print(row)
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=ExcelClass.get_column_letter(row.index(value) + 1),
                        row=global_workers_enbek_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=output_1)

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        threading.Thread(target=extra_enbek, args=()).start()
        
############################################################################################

