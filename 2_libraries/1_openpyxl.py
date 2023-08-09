########################################################################################################################
# TODO READ

import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook(filename='temp/data.xlsx')
worksheet = workbook.active  # workbook['Sheet1']
print(worksheet['A1'], type(worksheet['A1']))  # <Cell 'Sheet1'.A1> <class 'openpyxl.cell.cell.Cell'>
print(worksheet['A1'].value)

print(worksheet.cell(1, 1))
print(worksheet.cell(1, 1).value)

max_row = worksheet.max_row
max_column = worksheet.max_column

# vals = ["A1", "A2"]
# idx = 0
# for item in vals:
#     idx += 1
#     # ...


rows = []
for row_index in range(1, max_row + 1):
    row = []
    for column_index in range(1, max_column + 1):
        # cell = worksheet[f"{get_column_letter(column_index)}{row_index}"]
        print(get_column_letter(column_index))
        cell = worksheet.cell(row_index, column_index)
        value = cell.value
        row.append(value)
    rows.append(row)

########################################################################################################################


########################################################################################################################
# TODO CREATE

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active
new_worksheet.title = 'page 1'

rows = [[f"{column_index} {row_index}" for column_index in "ABCDEFGH"] for row_index in range(1, 10)]

# a1 = worksheet[f"A1"]
# a1.font = openpyxl.styles.Font(color="00FF6600", bold=True, sz=28)
# worksheet[f"A1"] = "Фамилия Имя"
# worksheet.merge_cells('B2:C6')
# top_left_cell = worksheet['B2']
# top_left_cell.value = "My Cell"
# thin = openpyxl.styles.Side(border_style="thin", color="000000")
# double = openpyxl.styles.Side(border_style="double", color="ff0000")
# top_left_cell.border = openpyxl.styles.Border(top=double, left=thin, right=thin, bottom=double)
# top_left_cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")
# top_left_cell.fill = fill = openpyxl.styles.GradientFill(stop=("000000", "FFFFFF"))
# top_left_cell.font = openpyxl.styles.Font(b=True, color="FF0000")
# top_left_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

# font = openpyxl.styles.Font(name='Tahoma', size=16, bold=True,
#                             italic=False, vertAlign=None, underline='none',
#                             strike=False, color='FF0000FF')
# worksheet['B2'].font = font
# worksheet['B2'] = 'Python'

index_row1 = 0
for row_index, row in enumerate(rows, 1):
    for column_index, value in enumerate(row, 1):
        # worksheet[f"{get_column_letter(column_index)}{row_index}"] = value
        new_worksheet.cell(row_index, column_index, value)

new_workbook.save("temp/new_data.xlsx")

########################################################################################################################
