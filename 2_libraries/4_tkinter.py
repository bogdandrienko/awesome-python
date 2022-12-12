#####################
# сборка в exe

# pip install pyinstaller # обязательно внутри нужного виртуального окружения
# pyinstaller --onefile --windowed ui_1_tkinter.py
# pyinstaller --onedir --windowed ui_1_tkinter.py

#####################





# 1) ссылка, количество, название - программа для скачки картинок (Tkinter)
# 2) Парсер погоды (PySide2)
# 3) Парсер валюты (Pyqt6)

# from tkinter import *  # коллизии имён
from tkinter import ttk
import tkinter
import requests


root = tkinter.Tk()
root.geometry("640x480")
frm = ttk.Frame(root, padding=10)

frm.grid()

ttk.Label(frm, text="Ссылка: ").grid(column=0, row=0)
url = ttk.Entry(frm)
url.grid(column=1, row=0)

ttk.Label(frm, text="Количество: ").grid(column=0, row=1)
count = ttk.Entry(frm)
count.grid(column=1, row=1)

ttk.Label(frm, text="Имя: ").grid(column=0, row=2)
name = ttk.Entry(frm)
name.grid(column=1, row=2)


def start():
    print("start")
    # print(url, type(url))
    # print(url.get(), type(url.get()))
    # url_for_download = str(url.get())   # https://picsum.photos/320/240/
    # count_for_download = int(count.get())
    # name_for_download = str(name.get())
    val1 = int(url.get())
    val2 = int(count.get())

    def summator(start_value: int, stop_value: int) -> int:
        sum_value = 0
        for i in range(start_value, stop_value + 1, 1):
            sum_value += i
        return sum_value

    value = summator(start_value=val1, stop_value=val2)
    print(value)
    # print(url_for_download)
    # print(count_for_download)
    # print(name_for_download)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/102.0.0.0 Safari/537.36'
    }
    # data = b"hello"
    # "hello".encode() => b"hello" | b"hello".decode() => "hello"

    # for i in range(1, count_for_download + 1):  # 1 2 3 ... 7
    #     data = requests.get(url=url_for_download, headers=headers).content
    #     with open(f"{name_for_download}({i}).jpg", "wb") as opened_file:
    #         opened_file.write(data)

    print("stop")


ttk.Button(frm, text="Старт", command=start).grid(column=1, row=3)

root.mainloop()




#####################
# сборка в exe

# pip install pyinstaller # обязательно внутри нужного виртуального окружения
# pyinstaller --onefile --windowed ui_1_tkinter.py
# pyinstaller --onedir --windowed ui_1_tkinter.py

#####################


# 1) ссылка, количество, название - программа для скачки картинок (Tkinter)
# 2) Парсер погоды (PySide2)
# 3) Парсер валюты (Pyqt6)

# from tkinter import *  # коллизии имён
from tkinter import ttk
import tkinter
import openpyxl

root = tkinter.Tk()
root.geometry("640x480")
frm = ttk.Frame(root, padding=10)

frm.grid()

ttk.Label(frm, text="Путь, где лежит файл: ").grid(column=0, row=0)
url = ttk.Entry(frm)
url.grid(column=1, row=0)
url.insert(0, "data/new.xlsx")

ttk.Label(frm, text="Колонки, для чтения: ").grid(column=0, row=1)
count = ttk.Entry(frm)
count.grid(column=1, row=1)
count.insert(0, "1, 4,7, 9")

ttk.Label(frm, text="Искомое: ").grid(column=0, row=2)
name = ttk.Entry(frm)
name.grid(column=1, row=2)


def start():
    print("start")

    url_for_download = str(url.get())
    # path = "data/new.xlsx"
    count_for_download = str(count.get()).split(sep=",")
    name_for_download = str(name.get())

    clear_count_for_download = []  # [1, 4, 7, 9]
    for i in count_for_download:  # ["1", " 4", "7 ", " 9"]
        int_value = int(i.strip())
        clear_count_for_download.append(int_value)

    # workbook = openpyxl.Workbook()  # TODO НОВЫЙ ФАЙЛ
    workbook = openpyxl.load_workbook(url_for_download)
    worksheet = workbook.active

    cell1 = worksheet.cell(1, 1).value
    print(cell1)

    rows1 = [  # rows
        [666, 2, 3],  # row [value, value, value]
        [66, 3, 4],
        [3, 4, 5],
    ]
    rows = []
    for row_index in range(1, 20 + 1):
        if row_index not in clear_count_for_download:  # 1, 3, 5, 9
            continue
        row = []
        for column_index in range(1, 10 + 1):
            cell = worksheet.cell(row_index, column_index)
            value = cell.value
            if value is None:
                value = 0
            # if value % 2 == 0:
            #     value = "Чётное"
            # else:
            #     value = "Нечётное"
            row.append(value)
        rows.append(row)
    print(rows)
    print("stop")


ttk.Label(frm, text="Количество совпадений: ").grid(column=2, row=3)
ttk.Button(frm, text="Старт", command=start).grid(column=1, row=3)

root.mainloop()

# путь и имя файла который нужно читать ("data/new.xlsx")
# перечисление колонок, которые нужно читать из файла ("1, 5,7, 9")
# элемент для поиска (+ вывод на экран количества совпадений)
# сортирует файл и перезаписывает в новый (по lambda)



#####################
# сборка в exe

pip install pyinstaller # обязательно внутри нужного виртуального окружения
pyinstaller --onefile --console withou_ui.py
pyinstaller --onedir --console withou_ui.py

#####################

val1 = int(input("Введите первое число: "))
val2 = int(input("Введите второе число: "))

print(f"Результат: {val1*val2}")

val3 = int(input("введите что-либо для закрытия: "))


########################################################################################################################

from tkinter import *

value = 0


def calc():
    global value

    print(f'chk_value: {chk_value.get()}')
    print(f'entr_value: {entr_value.get()}')

    if chk_value.get():
        str1 = entr_value.get()
        if str(str1).isdigit():
            value += int(str1)
            lbl.config(text=value)
        else:
            lbl.config(text="Вы ввели некорректное значение!")
    print('press')


window = Tk()
window.title("Welcome to LikeGeeks app")
window.geometry('350x200')

lbl = Label(window, text="Hello")
lbl.grid(column=0, row=0)

btn = Button(window, text="Click Me", command=calc)  # тут нужна ссылка на функцию (имя функции)
btn.grid(column=0, row=1)

entr_value = StringVar(value='1')
entr = Entry(textvariable=entr_value, )
entr.grid(column=1, row=0)

chk_value = IntVar()
chk = Checkbutton(text="Добавлять/не добавлять", variable=chk_value)
chk.grid(column=1, row=1)

window.mainloop()

#######################################################################################################################

# импорт всей библиотеки
import time

# импорт всех функций, классов и переменных из библиотеки ! Может вызвать коллизию имён !
# from tkinter import *

# импорт всей библиотеки с присовением псевдонима
# import tkinter as tk


# импорт отдельной функции из модуля(из нашего файла)
# from .calc import calc_3


# импортируем библиотеку для работы с окнами(интерфейсом) все библиотеки: tkinter pyqt5 pyside6


# import tkinter
# import tkinter.ttk as ttk
# root = tkinter.Tk()
# frm = ttk.Frame(root, padding=10)
# frm.grid()
# ttk.Label(frm, text="Hello World!").grid(column=0, row=0)
# ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=0)
# root.mainloop()

from tkinter import *
from tkinter import ttk

# инициализация инстанса - создание объекта ткинтер
root = Tk()

# создание главного окна
frm = ttk.Frame(root, padding=100)
root.title("Таймер с интерфейсом на Python")
root.geometry("640x480")
frm.grid()
# Мы пишем таймер

# часы
ttk.Label(frm, text="00").grid(column=0, row=0)

# двоеточие
ttk.Label(frm, text=":").grid(column=1, row=0)

# минуты
ttk.Label(frm, text="00").grid(column=2, row=0)

# двоеточие
ttk.Label(frm, text=":").grid(column=3, row=0)

# секунды
ttk.Label(frm, text="00").grid(column=4, row=0)

# кнопка стоп
Button(text="Hello").grid(column=1, row=1)

# кнопка старт
Button(text="Hello").grid(column=3, row=1)

# ttk.Label(frm, text="Any").grid(column=0, row=1)
# ttk.Label(frm, text="Alex").grid(column=1, row=1)
# ttk.Label(frm, text="Ivan").grid(column=2, row=1)

# 0    1    2     3    4      5
# 1   Any    :   Alex  :     Ivan



# ttk.Label(frm, text="Hello World!").grid(column=0, row=0)

# ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=0)


root.mainloop()

########################################################################################################################

# импорт всей библиотеки
import time
from threading import Thread

# pip install pyinstaller
# pyinstaller --onefile ui.py
# python -m PyInstaller --onefile ui.py

# импорт всех функций, классов и переменных из библиотеки ! Может вызвать коллизию имён !
# from tkinter import *

# импорт всей библиотеки с присовением псевдонима
# import tkinter as tk


# импорт отдельной функции из модуля(из нашего файла)
# from .calc import calc_3


# импортируем библиотеку для работы с окнами(интерфейсом) все библиотеки: tkinter pyqt5 pyside6


# import tkinter
# import tkinter.ttk as ttk
# root = tkinter.Tk()
# frm = ttk.Frame(root, padding=10)
# frm.grid()
# ttk.Label(frm, text="Hello World!").grid(column=0, row=0)
# ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=0)
# root.mainloop()

from tkinter import *
from tkinter import ttk

hours = 0
minutes = 0
seconds = 0

pause = True


def stop_timer():
    global pause

    pause = False


def reset_timer():
    global hours
    hours = 0
    global minutes
    minutes = 0
    global seconds
    seconds = 0
    hours_label.config(text=f"{hours}")
    minuts_label.config(text=f"{minutes}")
    seconds_label.config(text=f"{seconds}")


def start_timer():
    global pause

    pause = True

    global hours
    # hours = 0
    global minutes
    # minutes = 0
    global seconds
    # seconds = 0

    # код до цикла
    while pause:
        # seconds = seconds + 1
        seconds += 1

        if seconds > 59:
            minutes += 1
            seconds = 0

        if minutes > 59:
            hours += 1
            minutes = 0

        if hours > 23:
            seconds = 0
            minutes = 0
            hours = 0

        time.sleep(0.00001)

        hours_label.config(text=f"{hours}")
        minuts_label.config(text=f"{minutes}")
        seconds_label.config(text=f"{seconds}")
        print(f"{hours}:{minutes}" + ":" + str(seconds))


# определение(создание) функции
def start_new_thread():
    Thread(
        target=start_timer
    ).start()  # Используйте для вызова функции в отдельный поток, тогда остальная часть окна не
    # будет виснуть

# вызов функции
# start_new_thread()

# ссылка на функцию
# start_new_thread

# инициализация инстанса - создание объекта ткинтер
root = Tk()

# создание главного окна
frm = ttk.Frame(root, padding=100)
root.title("Таймер с интерфейсом на Python")
root.geometry("640x480")
frm.grid()
# Мы пишем таймер

# часы
hours_label = ttk.Label(frm, text="00")
hours_label.grid(column=0, row=0)

# двоеточие
ttk.Label(frm, text=":").grid(column=1, row=0)

# минуты
minuts_label = ttk.Label(frm, text="00")
minuts_label.grid(column=2, row=0)

# двоеточие
ttk.Label(frm, text=":").grid(column=3, row=0)

# секунды
seconds_label = ttk.Label(frm, text="00")
seconds_label.grid(column=4, row=0)


# кнопка стоп
Button(text="stop",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=stop_timer,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=0, row=1)

# кнопка сброс
Button(text="reset",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=reset_timer,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=1, row=1)

# кнопка старт
Button(text="start",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=start_new_thread,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=2, row=1)

# ttk.Label(frm, text="Any").grid(column=0, row=1)
# ttk.Label(frm, text="Alex").grid(column=1, row=1)
# ttk.Label(frm, text="Ivan").grid(column=2, row=1)

# 0    1    2     3    4      5
# 1   Any    :   Alex  :     Ivan


# ttk.Label(frm, text="Hello World!").grid(column=0, row=0)

# ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=0)

# print_to_console("something")

root.mainloop()

########################################################################################################################

import time
# from tkinter import *  # коллизия имён!!!
import tkinter
import threading


def get_long_time(tim: int) -> str:
    """
    # 1 - 01 , 9 - 09, 10 - 10
    """
    if tim < 10:
        return f"0{tim}"
    return f"{tim}"


is_play = True
slider_widget = None
label_widget = None

hr = 0
mn = 0
sec = 0


def start():
    global hr
    global mn
    global sec

    global is_play
    is_play = True
    while True:
        multiplay = 1
        global slider_widget
        if slider_widget is not None:
            multiplay = slider_widget.get()

        if is_play:
            if sec >= 59:
                if mn >= 59:
                    if hr >= 23:
                        hr = 0
                        mn = 0
                        sec = 0
                    else:
                        hr += 1
                        mn = 0
                        sec = 0
                else:
                    mn += 1
                    sec = 0
            else:
                sec += 1 * multiplay

            str1 = f"{get_long_time(hr)}:{get_long_time(mn)}:{get_long_time(sec)}"
            print(str1)
            label_widget.config(text=str1)

        time.sleep(1)


def pause():
    global is_play
    is_play = not is_play


def reset():
    global hr
    hr = 0
    global mn
    mn = 0
    global sec
    sec = 0
    label_widget.config(text="00:00:00")



def start_thread():
    new_thread = threading.Thread(target=start)
    new_thread.start()

    # with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
    #     executor.submit(start)


def create_ui():
    tkinter_ui_window = tkinter.Tk()
    label_time = tkinter.Label(tkinter_ui_window, text="Hello Python", fg="#eee", bg="#333")
    label_time.place(x=0, y=0)
    btn_start = tkinter.Button(tkinter_ui_window, text="start", fg='blue', command=start_thread)
    btn_start.place(x=0, y=50)
    btn_reset = tkinter.Button(tkinter_ui_window, text="reset", fg='yellow', command=reset)
    btn_reset.place(x=0, y=100)
    btn_pause = tkinter.Button(tkinter_ui_window, text="pause", fg='green', command=pause)
    btn_pause.place(x=0, y=150)
    slider = tkinter.Scale(tkinter_ui_window, from_=1, to=59, orient=tkinter.HORIZONTAL)
    slider.place(x=0, y=200)

    global slider_widget
    slider_widget = slider

    global label_widget
    label_widget = label_time

    print(slider.get())

    # lbl = Label(window, text="This is Label widget", fg='red', font=("Helvetica", 16))
    # lbl.place(x=60, y=50)
    # txtfld = Entry(window, bd=5)
    # txtfld.place(x=80, y=150)
    tkinter_ui_window.title('Hello Python')
    tkinter_ui_window.geometry("640x480+10+10")
    tkinter_ui_window.mainloop()


if __name__ == '__main__':
    create_ui()
    # start()

########################################################################################################################

from tkinter import *

value = 0


def calc():
    global value

    print(f'chk_value: {chk_value.get()}')
    print(f'entr_value: {entr_value.get()}')

    if chk_value.get():
        str1 = entr_value.get()
        if str(str1).isdigit():
            value += int(str1)
            lbl.config(text=value)
        else:
            lbl.config(text="Вы ввели некорректное значение!")
    print('press')


window = Tk()
window.title("Welcome to LikeGeeks app")
window.geometry('350x200')

lbl = Label(window, text="Hello")
lbl.grid(column=0, row=0)

btn = Button(window, text="Click Me", command=calc)  # тут нужна ссылка на функцию (имя функции)
btn.grid(column=0, row=1)

entr_value = StringVar(value='1')
entr = Entry(textvariable=entr_value, )
entr.grid(column=1, row=0)

chk_value = IntVar()
chk = Checkbutton(text="Добавлять/не добавлять", variable=chk_value)
chk.grid(column=1, row=1)

window.mainloop()


##############################################################################


class TkinterClass:
    @staticmethod
    def example():
        def play_func(data: str):
            global play
            play = False
            sleep(0.5)
            play = True

            def whiles():
                print(data)

            thread_render = Thread(target=whiles)
            thread_render.start()
            app.set_title("Завершено")

        class Application(tkinter.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.id = 0
                self.iid = 0
                self.root = root
                self.root.title("ожидание")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')
                self.master.minsize(1280, 720)
                self.master.maxsize(1280, 720)

                self.play_btn = tkinter.Button(self.root, text="Запустить", font="100", command=self.play_button)
                self.play_btn.grid(row=0, column=0, sticky=tkinter.W)
                self.stop_btn = tkinter.Button(self.root, text="Остановить", font="100", command=self.stop_button)
                self.stop_btn.grid(row=0, column=1, sticky=tkinter.W)
                self.quit_btn = tkinter.Button(self.root, text="Выход", font="100", command=self.quit_button)
                self.quit_btn.grid(row=0, column=2, sticky=tkinter.W)

                self.export_label = tkinter.Label(self.root, text="Видеофайл для анализа/ip для анализа", font="100")
                self.export_label.grid(row=1, column=0, sticky=tkinter.W)
                self.export_entry = tkinter.Entry(self.root, font="100")
                self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
                self.export_entry.insert(0, 'video.mp4')

                chk_state = tk.BooleanVar()
                chk_state.set(False)
                chk = ttk.Checkbutton(self.root, text='Выбрать', var=chk_state)
                chk_state.set(False)
                chk.grid(row=2, column=1)

                self.combo = ttk.Combobox(self.root)
                self.combo['values'] = (1, 2, 3, 4, 5, "Text")
                self.combo.current(1)
                self.combo.grid(row=3, column=1, sticky=tkinter.W)

                self.text = tkinter.Text(self.root, font="100")
                self.text.grid(row=4, column=0, sticky=tkinter.W)

            def play_button(self):
                self.set_title("в процессе")
                play_func(data="старт")

            def stop_button(self):
                self.set_title("пауза")
                global play
                play = False

            def quit_button(self):
                self.set_title("выход")
                global play
                play = False
                self.quit()

            def set_title(self, title: str):
                self.root.title(title)

        if __name__ == "__main__":
            play = False
            app = Application(tk.Tk())
            thread_main = Thread(target=app.root.mainloop())
            thread_main.start()

    @staticmethod
    def example_generate_list():
        class Application(tk.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.root = root
                self.initialize_user_interface()

            def initialize_user_interface(self):
                self.root.title("Попытка в приложение")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')

                # Define the different GUI widgets
                self.SurnameNumber_label = tk.Label(self.root, text="Фамилия:", font="100")
                self.SurnameNumber_entry = tk.Entry(self.root, font="100")
                self.SurnameNumber_label.grid(row=0, column=0, sticky=tk.W)
                self.SurnameNumber_entry.grid(row=0, column=0)

                self.NameName_label = tk.Label(self.root, text="Имя:", font="100")
                self.NameName_entry = tk.Entry(self.root, font="100")
                self.NameName_label.grid(row=1, column=0, sticky=tk.W)
                self.NameName_entry.grid(row=1, column=0)

                self.PatronymicName_label = tk.Label(self.root, text="Отчество:", font="100")
                self.PatronymicName_entry = tk.Entry(self.root, font="100")
                self.PatronymicName_label.grid(row=2, column=0, sticky=tk.W)
                self.PatronymicName_entry.grid(row=2, column=0)

                self.Extra_label = tk.Label(self.root, text="Дополнительно:", font="100")
                self.Extra_entry = tk.Entry(self.root, font="100")
                self.Extra_label.grid(row=3, column=0, sticky=tk.W)
                self.Extra_entry.grid(row=3, column=0)

                self.submit_button = tk.Button(self.root, text="Добавить", font="100", command=self.insert_data)
                self.submit_button.grid(row=2, column=1, sticky=tk.W)

                self.exit_button = tk.Button(self.root, text="Выход", font="100", command=self.quit)
                self.exit_button.grid(row=0, column=1, sticky=tk.W)

                # Set the treeview
                self.tree = ttk.Treeview(self.root, columns=('№', 'Фамилия:', 'Имя:', 'Отчество:', 'Дополнительно:'))

                # Set the heading (Attribute Names)
                self.tree.heading('#0', text='№')
                self.tree.heading('#1', text='Фамилия')
                self.tree.heading('#2', text='Имя')
                self.tree.heading('#3', text='Отчество')
                self.tree.heading('#4', text='Дополнительно')

                # Specify attributes of the columns (We want to stretch it!)
                self.tree.column('#0', stretch=tk.YES)
                self.tree.column('#1', stretch=tk.YES)
                self.tree.column('#2', stretch=tk.YES)
                self.tree.column('#3', stretch=tk.YES)
                self.tree.column('#4', stretch=tk.YES)

                self.tree.grid(row=3, columnspan=4, sticky='nsew')
                self.treeview = self.tree

                self.id = 0
                self.iid = 0

            def insert_data(self):
                self.treeview.insert('', 'end', iid=self.iid, text=str(self.id + 1),
                                     values=(self.SurnameNumber_entry.get(), self.NameName_entry.get(),
                                             self.PatronymicName_entry.get(),
                                             str(self.SurnameNumber_entry.get() + self.NameName_entry.get() +
                                                 self.PatronymicName_entry.get())))
                self.iid = self.iid + 1
                self.id = self.id + 1

        app = Application(tk.Tk())
        thread_main = Thread(target=app.root.mainloop())
        thread_main.start()


########################################################################################################################


from tkinter import *  # может произойти коллизия имён (перезапись предыдущего имени последующим)
from tkinter import ttk
import tkinter
from tkinter import messagebox

import openpyxl


def cancel():
    pass


# определение(создание) функции
def get_result():
    col_from_user = col_label.get()
    print(f'колонки: {col_from_user}')
    elem_from_user = elem_label.get()
    print(f'элемент: {elem_from_user}')
    ############################################################

    file_name = 'temp/sample_example_new.xlsx'
    workbook = openpyxl.load_workbook(file_name)

    # берём активную страницу из рабочей книги
    worksheet = workbook.active

    # берём последнюю строку в excel-файле
    max_row = worksheet.max_row
    print(max_row)
    # берём последнюю колонку в excel-файле
    max_column = worksheet.max_column
    print(max_column)

    # записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
    external_array = []

    # наполнение внешнего массива
    for row in range(1, max_row + 1):
        internal_array = []

        # наполнение внутреннего массива
        for col in range(1, max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ''
                # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
                # break  # (прекратить) оператор позволяет полностью остановить цикл
            internal_array.append(value)
        # завершение наполнение внутреннего массива

        if len(internal_array) < 1:
            continue

        external_array.append(internal_array)
    # завершение наполнение внешнего массива

    print(external_array)

    # '1, 2, 6' -> [1, 2, 6]
    col_list = []
    print(col_from_user)
    print(type(col_from_user))
    for i in col_from_user.split(","):  # метод, который вызывается на строках (режет строку на массив)
        # "1, 2, 3 , 5".split(",") -> ["1", " 2", " 3 ", " 5"]
        i = str(i).strip()  # " 1" / "2 " -> "1" / "2"
        i = int(i)  # "1" -> 1
        col_list.append(i)
        print(i)
        print(type(i))
    print(col_list)

    # elem_from_user
    count = 0
    for i_new in external_array:
        for j_new in i_new:
            if elem_from_user == j_new:
                count = count + 1

    ###########################################################

    # тут мы уже должны посчитать количество вхождений
    result = count
    result_label.config(text=f"{result}")


# инициализация инстанса - создание объекта ткинтер
root = Tk()

# создание главного окна
frm = ttk.Frame(root, padding=100)
root.title("Анализ данных")
root.geometry("640x480")
frm.grid()

# результат
result_label = ttk.Label(frm, text="Количество вхождений: ")
result_label.grid(column=4, row=0)

# колонки
col_label_var = tkinter.StringVar()  # переменная, которая хранит колонки
col_label = ttk.Entry(textvariable=col_label_var)
col_label.grid(column=1, row=3)

# элемент
elem_label_var = tkinter.StringVar()  # переменная, которая хранит элемент
elem_label = ttk.Entry(textvariable=elem_label_var)
elem_label.grid(column=5, row=3)

# кнопка стоп
Button(text="stop",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=cancel,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=0, row=5)

# кнопка старт
Button(text="start",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=get_result,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=2, row=5)

root.mainloop()

str1 = "124124|14121|1411234".split(sep="|")  # -> ["124124", "14121", "1411234"]

########################################################################################################################


